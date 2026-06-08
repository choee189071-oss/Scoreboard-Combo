"""
CreditScope source loader.

This connector wraps the row/column mapping logic in engine.mapping_engine and
emits the unified source-candidate schema used by engine.data_sourcing_engine.
Most rows are raw source fields; S&P support-tab methodology metrics are also
allowed when they avoid ambiguous workbook units or denominators.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Optional, Union

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

from engine.data_sourcing_engine import (
    CANDIDATE_COLUMNS,
    mapping_report_to_source_candidates,
    normalize_source_candidates,
)
from engine.mapping_engine import clean_numeric_value, map_creditscope_workbook, map_uploaded_file, normalize_label


def _uploaded_name(uploaded_file: Any) -> str:
    if isinstance(uploaded_file, (str, Path)):
        return Path(uploaded_file).name
    return str(getattr(uploaded_file, "name", "") or "")


def _is_excel(uploaded_file: Any) -> bool:
    return Path(_uploaded_name(uploaded_file)).suffix.lower() in {".xlsx", ".xls"}


def _seek_start(uploaded_file: Any) -> None:
    if not isinstance(uploaded_file, (str, Path)) and hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)


def _open_workbook(uploaded_file: Any):
    _seek_start(uploaded_file)
    return load_workbook(uploaded_file, read_only=True, data_only=True)


def _sheet_by_normalized_name(workbook: Any, sheet_name: str):
    target = normalize_label(sheet_name)
    for name in workbook.sheetnames:
        if normalize_label(name) == target:
            return workbook[name]
    return None


def _find_label_row(ws: Any, aliases: Iterable[str], label_col: int = 2) -> tuple[int, str] | tuple[None, None]:
    alias_set = {normalize_label(alias) for alias in aliases}
    for row_idx in range(1, ws.max_row + 1):
        label = ws.cell(row_idx, label_col).value
        if normalize_label(label) in alias_set:
            return row_idx, str(label).strip()
    return None, None


def _read_period_series(
    ws: Any,
    *,
    aliases: Iterable[str],
    label_col: int = 2,
    header_row: int = 7,
    start_col: int = 3,
    max_periods: int = 3,
    absolute: bool = False,
) -> tuple[list[float], str, str] | tuple[None, None, None]:
    row_idx, label = _find_label_row(ws, aliases, label_col=label_col)
    if row_idx is None:
        return None, None, None

    values: list[float] = []
    first_col: int | None = None
    last_col: int | None = None
    for col_idx in range(start_col, start_col + max_periods):
        period_header = ws.cell(header_row, col_idx).value
        if period_header is None:
            continue
        value = clean_numeric_value(ws.cell(row_idx, col_idx).value)
        if value is None or not isinstance(value, (int, float)):
            continue
        numeric = abs(float(value)) if absolute else float(value)
        values.append(numeric)
        first_col = col_idx if first_col is None else first_col
        last_col = col_idx

    if not values or first_col is None or last_col is None:
        return None, None, None

    cell_range = f"{ws.title}!{get_column_letter(first_col)}{row_idx}:{get_column_letter(last_col)}{row_idx}"
    return values, cell_range, label


def _read_single_metric(
    ws: Any,
    *,
    aliases: Iterable[str],
    label_col: int = 2,
    value_col: int = 6,
) -> tuple[float, str, str] | tuple[None, None, None]:
    row_idx, label = _find_label_row(ws, aliases, label_col=label_col)
    if row_idx is None:
        return None, None, None

    value = clean_numeric_value(ws.cell(row_idx, value_col).value)
    if value is None or not isinstance(value, (int, float)):
        return None, None, None

    source_cell = f"{ws.title}!{get_column_letter(value_col)}{row_idx}"
    return float(value), source_cell, label


def _supplemental_match_row(
    *,
    sheet_name: str,
    field_name: str,
    value: Any,
    source_cell: str,
    source_label: str,
    notes: str,
    match_method: str = "sp_local_government_supplemental_tab",
    status: str = "source_review",
) -> dict[str, Any]:
    return {
        "sheet_name": sheet_name,
        "field_name": field_name,
        "source_name": "CreditScope",
        "source_type": "Upload",
        "matched_column": source_cell,
        "matched_label": source_label,
        "match_method": match_method,
        "confidence": 0.99,
        "value": value,
        "notes": notes,
        "status": status,
    }


def _load_sp_local_gov_supplemental_report(uploaded_file: Any) -> pd.DataFrame:
    """
    Extract S&P local government inputs from scorecard support tabs.

    These tabs contain raw multi-year inputs used by S&P's local-government
    operating-result and reserve metrics. They also contain direct methodology
    metric values for debt and pension measures where raw totals often have
    unit, geography, or adjustment ambiguity.
    """
    rows: list[dict[str, Any]] = []
    workbook = None
    try:
        workbook = _open_workbook(uploaded_file)

        financial_ws = _sheet_by_normalized_name(workbook, "Financial Performance")
        if financial_ws is not None:
            financial_specs = [
                (
                    "governmental_revenue",
                    ["Governmental Revenues"],
                    False,
                    "Three-period governmental revenue series from S&P Financial Performance tab.",
                ),
                (
                    "governmental_expense",
                    ["Governmental Expenses"],
                    True,
                    "Three-period governmental expense series from S&P Financial Performance tab; signs are normalized to positive expenses.",
                ),
                (
                    "operating_transfers",
                    ["Transfers"],
                    False,
                    "Three-period transfer series from S&P Financial Performance tab; signed values are retained.",
                ),
            ]
            for field_name, aliases, absolute, notes in financial_specs:
                value, source_cell, source_label = _read_period_series(
                    financial_ws,
                    aliases=aliases,
                    absolute=absolute,
                )
                if value is None:
                    continue
                rows.append(
                    _supplemental_match_row(
                        sheet_name=financial_ws.title,
                        field_name=field_name,
                        value=value,
                        source_cell=source_cell,
                        source_label=source_label,
                        notes=notes,
                    )
                )

        reserves_ws = _sheet_by_normalized_name(workbook, "Reserves and Liquidity")
        if reserves_ws is not None:
            reserve_specs = [
                (
                    "committed_fund_balance",
                    ["Committed to"],
                    "Three-period committed fund balance series from S&P Reserves and Liquidity tab.",
                ),
                (
                    "assigned_fund_balance",
                    ["Assigned"],
                    "Three-period assigned fund balance series from S&P Reserves and Liquidity tab.",
                ),
                (
                    "unassigned_fund_balance",
                    ["Unassigned"],
                    "Three-period unassigned fund balance series from S&P Reserves and Liquidity tab.",
                ),
                (
                    "reserve_revenue",
                    ["Governmental Revenues"],
                    "Three-period reserve denominator revenue series from S&P Reserves and Liquidity tab.",
                ),
            ]
            for field_name, aliases, notes in reserve_specs:
                value, source_cell, source_label = _read_period_series(reserves_ws, aliases=aliases)
                if value is None:
                    continue
                rows.append(
                    _supplemental_match_row(
                        sheet_name=reserves_ws.title,
                        field_name=field_name,
                        value=value,
                        source_cell=source_cell,
                        source_label=source_label,
                        notes=notes,
                    )
                )

        economy_ws = _sheet_by_normalized_name(workbook, "Economy")
        if economy_ws is not None:
            economy_metric_specs = [
                (
                    "gdp_per_capita_ratio",
                    ["Per Capita"],
                    "Direct S&P local-government real GCP per-capita ratio from Economy support tab.",
                ),
                (
                    "personal_income_ratio",
                    ["Per Capita Personal Income"],
                    "Direct S&P local-government PCPI ratio from Economy support tab.",
                ),
            ]
            for field_name, aliases, notes in economy_metric_specs:
                value, source_cell, source_label = _read_single_metric(
                    economy_ws,
                    aliases=aliases,
                    value_col=5,
                )
                if value is None:
                    continue
                rows.append(
                    _supplemental_match_row(
                        sheet_name=economy_ws.title,
                        field_name=field_name,
                        value=value,
                        source_cell=source_cell,
                        source_label=source_label,
                        notes=notes,
                        match_method="sp_local_government_direct_metric",
                        status="ready",
                    )
                )

        debt_ws = _sheet_by_normalized_name(workbook, "Debt and Liabilities")
        if debt_ws is not None:
            debt_metric_specs = [
                (
                    "fixed_cost_burden_ratio",
                    ["Current cost for debt service and liabilities"],
                    "Direct S&P local-government fixed-cost burden metric from Debt and Liabilities support tab.",
                ),
                (
                    "net_direct_debt_per_capita",
                    ["Net direct debt per capita"],
                    "Direct S&P local-government net direct debt per capita metric from Debt and Liabilities support tab.",
                ),
                (
                    "npl_per_capita",
                    ["Net pension liability (NPL) per capita", "Net pension liability per capita"],
                    "Direct S&P local-government NPL per capita metric from Debt and Liabilities support tab.",
                ),
            ]
            for field_name, aliases, notes in debt_metric_specs:
                value, source_cell, source_label = _read_single_metric(debt_ws, aliases=aliases)
                if value is None:
                    continue
                rows.append(
                    _supplemental_match_row(
                        sheet_name=debt_ws.title,
                        field_name=field_name,
                        value=value,
                        source_cell=source_cell,
                        source_label=source_label,
                        notes=notes,
                        match_method="sp_local_government_direct_metric",
                        status="ready",
                    )
                )
    except Exception:
        return pd.DataFrame()
    finally:
        if workbook is not None:
            workbook.close()
        _seek_start(uploaded_file)

    return pd.DataFrame(rows)


def load_creditscope_source_candidates(
    uploaded_file: Any,
    *,
    mapping_path: Union[str, Path] = "config/field_mapping.csv",
    row_mapping_path: Union[str, Path] = "config/creditscope_row_mapping.csv",
    sheet_name: Optional[str] = None,
    value_col: int = 2,
    required_fields: Optional[Iterable[str]] = None,
    fuzzy_threshold: float = 0.92,
) -> dict[str, Any]:
    """
    Load a CreditScope CSV/XLSX file as standardized source candidates.

    Returns a dict with:
      - issuer_data: mapped field/value dict
      - match_report: human-readable mapping report
      - source_candidates: normalized candidate rows
    """
    uploaded_name = _uploaded_name(uploaded_file)

    if _is_excel(uploaded_file):
        _seek_start(uploaded_file)
        issuer_data, match_report = map_creditscope_workbook(
            uploaded_file=uploaded_file,
            mapping_path=mapping_path,
            row_mapping_path=row_mapping_path,
            sheet_name=sheet_name,
            value_col=value_col,
            fuzzy_threshold=fuzzy_threshold,
        )
        if not issuer_data:
            _seek_start(uploaded_file)
            issuer_data, match_report = map_uploaded_file(
                uploaded_file=uploaded_file,
                source_name="CreditScope",
                mapping_path=mapping_path,
                fuzzy_threshold=0.82,
            )
        _seek_start(uploaded_file)
        supplemental_report = _load_sp_local_gov_supplemental_report(uploaded_file)
        if not supplemental_report.empty:
            for _, supplemental_row in supplemental_report.iterrows():
                value = supplemental_row.get("value")
                if value is not None:
                    issuer_data[str(supplemental_row["field_name"])] = value
            match_report = pd.concat([supplemental_report, match_report], ignore_index=True)
    else:
        _seek_start(uploaded_file)
        issuer_data, match_report = map_uploaded_file(
            uploaded_file=uploaded_file,
            source_name="CreditScope",
            mapping_path=mapping_path,
            fuzzy_threshold=0.82,
        )

    report = match_report.copy() if match_report is not None else pd.DataFrame()
    if not report.empty:
        if "uploaded_file" not in report.columns:
            report.insert(0, "uploaded_file", uploaded_name)
        else:
            report["uploaded_file"] = report["uploaded_file"].replace("", uploaded_name).fillna(uploaded_name)
        if "source_type" not in report.columns:
            report["source_type"] = "Upload"

    candidates = mapping_report_to_source_candidates(report, uploaded_file=uploaded_name)
    if required_fields is not None and not candidates.empty:
        required = {str(field).strip() for field in required_fields if str(field).strip()}
        candidates = candidates[candidates["field_name"].isin(required)].reset_index(drop=True)

    return {
        "issuer_data": issuer_data,
        "match_report": report,
        "source_candidates": normalize_source_candidates(candidates)
        if not candidates.empty
        else pd.DataFrame(columns=CANDIDATE_COLUMNS),
    }
