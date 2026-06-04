from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from utils.ui_helpers import (
    action_panel,
    current_context_card,
    init_state,
    page_header,
    readiness_action,
    selected_source_report,
    source_readiness_counts,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    from engine.calculator_engine import load_formula_library, parse_required_fields
    from engine.factor_engine import load_factor_template
    from engine.mapping_engine import map_uploaded_file
    from engine.data_sourcing_engine import (
        mapping_report_to_source_candidates,
        manual_data_to_source_candidates,
        run_data_sourcing_pipeline,
    )
    from connectors.census_api import (
        CensusApiError,
        fetch_census_source_candidates,
        supported_census_candidate_fields,
    )
    from connectors.bea_api import (
        BeaApiError,
        fetch_bea_source_candidates,
        supported_bea_candidate_fields,
    )
    from connectors.creditscope_loader import load_creditscope_source_candidates
except Exception as exc:  # pragma: no cover - Streamlit display path
    st.set_page_config(page_title="Data Mapping", layout="wide")
    st.error("Could not import mapping/calculator engines.")
    st.exception(exc)
    st.stop()


st.set_page_config(page_title="Data Mapping", layout="wide")
init_state()
page_header(
    "Data Mapping",
    "Build one canonical issuer_data dictionary from uploaded files, API candidates, and manual source-pending fields.",
    "data_mapping",
)
current_context_card()

methodology_id = st.session_state.get("methodology_id", "moodys_ccd_go")
st.session_state.setdefault("api_source_candidates", {})
st.session_state.setdefault("api_source_reports", {})
if st.session_state.get("api_source_methodology_id") != methodology_id:
    st.session_state["api_source_candidates"] = {}
    st.session_state["api_source_reports"] = {}
    st.session_state["api_source_methodology_id"] = methodology_id


SOURCE_SESSION_KEYS = {
    "issuer_data",
    "source_report",
    "source_candidates",
    "source_readiness_summary",
    "source_match_reports",
    "uploaded_sources",
    "api_source_candidates",
    "api_source_reports",
}
SOURCE_WIDGET_PREFIXES = (
    "upload_",
    "sheet_",
    "allow_mismatched_sheet_",
    "worksheet_override_",
    "manual_source_editor_",
)


def _reset_source_session() -> None:
    """Clear source-selection state and source-input widgets for a clean rerun."""
    for key in SOURCE_SESSION_KEYS:
        st.session_state.pop(key, None)
    for key in list(st.session_state.keys()):
        if any(str(key).startswith(prefix) for prefix in SOURCE_WIDGET_PREFIXES):
            st.session_state.pop(key, None)
    st.session_state["api_source_candidates"] = {}
    st.session_state["api_source_reports"] = {}
    st.session_state["api_source_methodology_id"] = methodology_id
    st.session_state["source_reset_notice"] = "Source session reset. Upload/select sources again before saving issuer_data."


def _default_value_for_field(field_name: str) -> Any:
    defaults = {
        "population": 100000.0,
        "population_us": 330000000.0,
        "full_value": 12000000000.0,
        "assessed_value": 10000000000.0,
        "assessed_value_prior": 9500000000.0,
        "operating_revenue": 100000000.0,
        "operating_expense": 98000000.0,
        "governmental_revenue": 100000000.0,
        "governmental_expense": 98000000.0,
        "general_fund_balance": 25000000.0,
        "available_fund_balance": 25000000.0,
        "cash_balance": 30000000.0,
        "cash_and_investments": 30000000.0,
        "unrestricted_reserves": 30000000.0,
        "net_direct_debt": 20000000.0,
        "direct_debt": 20000000.0,
        "total_go_debt": 20000000.0,
        "net_pension_liability": 35000000.0,
        "mads": 6000000.0,
        "mads_requirement": 6000000.0,
        "issuer_mfi": 120000.0,
        "us_mfi": 100000.0,
        "mhi_adjusted_rpp": 120000.0,
        "us_median_income": 80000.0,
        "county_ebi": 120.0,
        "us_ebi": 100.0,
        "county_gdp_current": 105.0,
        "county_gdp_prior": 100.0,
        "us_gdp_current": 103.0,
        "us_gdp_prior": 100.0,
    }
    return defaults.get(field_name, None)


def _display_df(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize mixed object columns for Streamlit/Arrow display only."""
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df

    def _display_value(value: Any) -> str:
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        return str(value)

    display = df.copy()
    for col in display.columns:
        if display[col].dtype == "object":
            display[col] = display[col].map(_display_value)
    return display


def _mapping_report_display(df: pd.DataFrame) -> pd.DataFrame:
    """Make upload match diagnostics readable for end users."""
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df

    display = df.copy()
    match_method = display.get("match_method", pd.Series("", index=display.index)).fillna("").astype(str)
    matched_column = display.get("matched_column", pd.Series("", index=display.index))
    confidence = pd.to_numeric(display.get("confidence", pd.Series(0.0, index=display.index)), errors="coerce").fillna(0.0)

    found = ~match_method.eq("not_found") & matched_column.notna() & matched_column.astype(str).str.strip().ne("")
    display.insert(
        0,
        "mapping_status",
        [
            "mapped" if is_found and conf >= 0.78 else "review" if is_found else "not_found_in_this_file"
            for is_found, conf in zip(found.tolist(), confidence.tolist())
        ],
    )

    for col in ["source_type", "matched_column", "matched_label", "value", "notes"]:
        if col in display.columns:
            display[col] = display[col].where(display[col].notna(), "")
    if "source_type" in display.columns:
        display["source_type"] = display["source_type"].replace("", "upload")
    if "notes" in display.columns:
        display.loc[~found, "notes"] = display.loc[~found, "notes"].replace(
            "",
            "Not found in this upload. This is expected for fields supplied by Census, BEA, OS, ACFR, or manual input.",
        )
    return display


def _required_fields_for_methodology(methodology_id: str) -> pd.DataFrame:
    formulas = load_formula_library("config/formula_library.csv")
    template = load_factor_template(methodology_id, templates_dir="templates")
    formula_ids = set(template["formula_id"].dropna().astype(str))
    try:
        thresholds = pd.read_csv("config/scoring_thresholds.csv")
        related_thresholds = thresholds[thresholds["methodology_id"].astype(str) == str(methodology_id)]
        secondary_ids = set(related_thresholds["secondary_formula_id"].dropna().astype(str))
        secondary_ids.discard("")
        secondary_ids.discard("nan")
        formula_ids |= secondary_ids
    except Exception:
        pass
    rows: List[Dict[str, Any]] = []
    for _, formula in formulas[formulas["formula_id"].astype(str).isin(formula_ids)].iterrows():
        for field in parse_required_fields(formula.get("required_data", "")):
            if field == "manual":
                continue
            rows.append(
                {
                    "field_name": field,
                    "formula_id": formula["formula_id"],
                    "formula_name": formula["formula_name"],
                    "category": formula.get("category", ""),
                }
            )
    if not rows:
        return pd.DataFrame(columns=["field_name", "used_by", "category"])
    df = pd.DataFrame(rows)
    return (
        df.groupby("field_name", as_index=False)
        .agg(
            used_by=("formula_id", lambda x: "; ".join(sorted(set(map(str, x))))),
            category=("category", lambda x: "; ".join(sorted(set(str(v) for v in x if str(v))))),
        )
        .sort_values("field_name")
        .reset_index(drop=True)
    )


def _clean_edited_values(df: pd.DataFrame) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for _, row in df.iterrows():
        key = str(row.get("field_name", "")).strip()
        raw_value = row.get("value")
        if not key or raw_value in [None, ""]:
            continue
        try:
            data[key] = float(raw_value)
        except Exception:
            data[key] = raw_value
    return data


def _is_excel_upload(uploaded_file: Any) -> bool:
    name = str(getattr(uploaded_file, "name", "") or uploaded_file)
    return Path(name).suffix.lower() in {".xlsx", ".xls"}


def _excel_sheet_names(uploaded_file: Any) -> List[str]:
    if not _is_excel_upload(uploaded_file):
        return []
    workbook = None
    try:
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        return list(workbook.sheetnames)
    except Exception:
        return []
    finally:
        if workbook is not None:
            workbook.close()
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)


def _name_tokens(value: str) -> set[str]:
    ignored = {
        "scorecard",
        "moodys",
        "moody",
        "higher",
        "local",
        "gov",
        "water",
        "sewer",
        "utility",
        "xlsx",
        "xls",
        "2023",
        "2024",
        "2025",
        "2026",
        "fin",
        "ccd",
        "go",
        "sp",
        "k12",
    }
    return {
        token
        for token in re.findall(r"[a-z0-9]+", str(value).lower())
        if len(token) > 2 and token not in ignored
    }


def _all_name_tokens(value: str) -> set[str]:
    return {token for token in re.findall(r"[a-z0-9]+", str(value).lower()) if len(token) > 2}


def _sheet_match_score(sheet_name: str, uploaded_name: str) -> int:
    upload_tokens = _name_tokens(uploaded_name)
    if not upload_tokens:
        return 0
    sheet_tokens = _name_tokens(sheet_name)
    return len(upload_tokens & sheet_tokens)


def _sheet_raw_hint_score(sheet_name: str) -> int:
    lowered = str(sheet_name).lower()
    score = 0
    if any(token in lowered for token in ["fin", "raw", "creditscope", "credit scope", "financial"]):
        score += 2
    if any(token in lowered for token in ["scorecard", "public", "summary", "validation"]):
        score -= 3
    if "backup" in lowered:
        score -= 1
    return score


def _is_likely_raw_sheet(sheet_name: str, uploaded_name: str) -> bool:
    return _sheet_match_score(sheet_name, uploaded_name) > 0 and _sheet_raw_hint_score(sheet_name) >= 0


def _is_generic_raw_sheet(sheet_name: str) -> bool:
    generic_tokens = {
        "credit",
        "creditscope",
        "data",
        "financial",
        "finance",
        "raw",
        "scope",
        "source",
        "template",
    }
    tokens = _all_name_tokens(sheet_name)
    return bool(tokens) and tokens <= generic_tokens and _sheet_raw_hint_score(sheet_name) > 0


def _is_sheet_mismatch(sheet_name: str | None, uploaded_name: str) -> bool:
    """Return True when a named workbook sheet clearly does not match the uploaded file."""
    if not sheet_name:
        return False
    upload_tokens = _name_tokens(uploaded_name)
    if not upload_tokens:
        return False
    return _sheet_match_score(sheet_name, uploaded_name) == 0


def _preferred_sheet_index(sheet_names: List[str], uploaded_name: str) -> int | None:
    if not sheet_names:
        return None
    scored = [
        (
            _sheet_match_score(sheet, uploaded_name),
            _sheet_raw_hint_score(sheet),
            -idx,
        )
        for idx, sheet in enumerate(sheet_names)
        if _is_likely_raw_sheet(sheet, uploaded_name)
    ]
    if not scored:
        generic_scored = [
            (
                _sheet_raw_hint_score(sheet),
                -idx,
            )
            for idx, sheet in enumerate(sheet_names)
            if _is_generic_raw_sheet(sheet)
        ]
        if len(generic_scored) != 1:
            return None
        return -generic_scored[0][1]
    best_idx = -max(scored)[2]
    return max(0, min(best_idx, len(sheet_names) - 1))


try:
    required_fields = _required_fields_for_methodology(methodology_id)
except Exception as exc:
    st.warning(f"Could not load required fields for {methodology_id}: {exc}")
    required_fields = pd.DataFrame(columns=["field_name", "used_by", "category"])

required_field_names = required_fields["field_name"].dropna().astype(str).str.strip().tolist()


reset_notice = st.session_state.pop("source_reset_notice", None)
if reset_notice:
    st.success(reset_notice)

existing_source_report = st.session_state.get("source_report", pd.DataFrame())
readiness_action(existing_source_report)
existing_counts = source_readiness_counts(existing_source_report)
if existing_counts:
    summary_cols = st.columns(4)
    summary_cols[0].metric("Independent Ready", existing_counts.get("independent_ready", 0))
    summary_cols[1].metric("Source Pending", existing_counts.get("source_pending", 0))
    summary_cols[2].metric("Needs Review", existing_counts.get("needs_review", 0))
    summary_cols[3].metric("Missing", existing_counts.get("missing", 0))

reset_col, reset_note_col = st.columns([1, 4])
with reset_col:
    if st.button("Reset source session"):
        _reset_source_session()
        st.rerun()
with reset_note_col:
    st.caption("Use reset before switching workbook, worksheet, issuer, or methodology source inputs.")

st.divider()
st.subheader("Source Inputs")
st.caption("Upload one raw workbook when available, then add API candidates for demographic and economic fields.")

source_options = {
    "creditscope": ("CreditScope", "CreditScope workbook"),
    "ipeds": ("IPEDS_Excel", "IPEDS Excel"),
    "os": ("OS", "Official Statement table extract"),
    "acfr": ("ACFR", "ACFR / Audit table extract"),
}

mapped_candidate_frames: List[pd.DataFrame] = []
match_reports: List[pd.DataFrame] = []
cols = st.columns(4)
for i, (key, (source_name, label)) in enumerate(source_options.items()):
    with cols[i]:
        uploaded = st.file_uploader(label, type=["csv", "xlsx", "xls"], key=f"upload_{key}")
        if uploaded is None:
            continue
        selected_sheet_name = None
        if source_name == "CreditScope" and _is_excel_upload(uploaded):
            sheet_names = _excel_sheet_names(uploaded)
            if sheet_names:
                default_sheet_idx = _preferred_sheet_index(sheet_names, uploaded.name)
                if default_sheet_idx is not None:
                    selected_sheet_name = sheet_names[default_sheet_idx]
                    st.success(f"Auto-selected CreditScope worksheet: {selected_sheet_name}")
                    with st.expander("Advanced worksheet override", expanded=False):
                        use_override = st.checkbox(
                            "Use a different worksheet",
                            value=False,
                            key=f"worksheet_override_{key}_{uploaded.name}",
                        )
                        if use_override:
                            override_sheet_name = st.selectbox(
                                "CreditScope worksheet",
                                options=sheet_names,
                                index=default_sheet_idx,
                                key=f"sheet_{key}_{uploaded.name}",
                            )
                            if _is_sheet_mismatch(override_sheet_name, uploaded.name):
                                st.warning(
                                    "Selected worksheet name does not match the uploaded file name. "
                                    "Only use this for generic workbook sheet names."
                                )
                                allow_mismatch = st.checkbox(
                                    "Allow mismatched worksheet anyway",
                                    value=False,
                                    key=f"allow_mismatched_sheet_{key}_{uploaded.name}",
                                    help="Use only for workbooks whose sheet names are generic and cannot match the issuer file name.",
                                )
                                if not allow_mismatch:
                                    continue
                            selected_sheet_name = override_sheet_name
                else:
                    st.error(
                        "No matching CreditScope raw worksheet found. Upload a workbook with a matching raw/FIN sheet, "
                        "or use this workbook only for official scorecard validation."
                    )
                    st.caption(f"Detected worksheets: {', '.join(sheet_names)}")
                    with st.expander("Advanced worksheet override", expanded=False):
                        allow_mismatch = st.checkbox(
                            "Map a non-matching worksheet anyway",
                            value=False,
                            key=f"allow_mismatched_sheet_{key}_{uploaded.name}",
                            help="Use only when the workbook sheet name is generic but the sheet is truly the issuer raw data.",
                        )
                        if allow_mismatch:
                            selected_sheet_name = st.selectbox(
                                "CreditScope worksheet",
                                options=sheet_names,
                                index=0,
                                key=f"sheet_{key}_{uploaded.name}",
                            )
                            st.warning(f"Override selected: {selected_sheet_name}")
                        else:
                            continue
        try:
            if source_name == "CreditScope":
                loader_output = load_creditscope_source_candidates(
                    uploaded_file=uploaded,
                    mapping_path="config/field_mapping.csv",
                    row_mapping_path="config/creditscope_row_mapping.csv",
                    sheet_name=selected_sheet_name,
                    value_col=2,
                    required_fields=required_field_names,
                )
                source_data = loader_output["issuer_data"]
                report = loader_output["match_report"]
                source_candidates = loader_output["source_candidates"]
            else:
                source_data, report = map_uploaded_file(
                    uploaded_file=uploaded,
                    source_name=source_name,
                    mapping_path="config/field_mapping.csv",
                )
                source_candidates = mapping_report_to_source_candidates(report, uploaded_file=uploaded.name)
            st.session_state["uploaded_sources"][key] = uploaded.name
            st.success(f"{len(source_data)} fields mapped")
            if not report.empty:
                if "uploaded_file" not in report.columns:
                    report.insert(0, "uploaded_file", uploaded.name)
                match_reports.append(report)
                mapped_candidate_frames.append(source_candidates)
        except Exception as exc:
            st.error(f"Could not map {uploaded.name}.")
            st.exception(exc)

if match_reports:
    with st.expander("Source mapping reports", expanded=True):
        combined_report = pd.concat(match_reports, ignore_index=True)
        report_display = _mapping_report_display(combined_report)
        status_counts = report_display["mapping_status"].value_counts().to_dict() if "mapping_status" in report_display else {}
        report_cols = st.columns(3)
        report_cols[0].metric("Mapped From Upload", int(status_counts.get("mapped", 0)))
        report_cols[1].metric("Needs Review", int(status_counts.get("review", 0)))
        report_cols[2].metric("Not In This File", int(status_counts.get("not_found_in_this_file", 0)))
        st.caption(
            "This is an upload diagnostic, not the final missing-field list. "
            "Rows marked Not In This File may still be filled by Census, BEA, OS, ACFR, or the manual fallback editor."
        )
        mapped_rows = report_display[report_display["mapping_status"].isin(["mapped", "review"])]
        not_found_rows = report_display[report_display["mapping_status"].eq("not_found_in_this_file")]
        mapped_tab, not_found_tab, all_tab = st.tabs(["Mapped / Review", "Not In This File", "All Diagnostics"])
        with mapped_tab:
            st.dataframe(_display_df(mapped_rows), width="stretch", hide_index=True) if not mapped_rows.empty else st.info(
                "No fields were mapped from the uploaded file."
            )
        with not_found_tab:
            st.dataframe(_display_df(not_found_rows), width="stretch", hide_index=True) if not not_found_rows.empty else st.info(
                "No upload fields were missing from this file."
            )
        with all_tab:
            st.dataframe(_display_df(report_display), width="stretch", hide_index=True)

st.divider()
st.subheader("API Candidates")
st.caption("Census and BEA fetch raw source values only. Formula calculations remain in calculator_engine.")

census_default_fields = [field for field in required_field_names if field in supported_census_candidate_fields()]
with st.expander("Census ACS", expanded=False):
    geo_cols = st.columns(4)
    with geo_cols[0]:
        census_year = st.number_input("ACS year", min_value=2009, max_value=2026, value=2024, step=1)
    with geo_cols[1]:
        state_fips = st.text_input("State FIPS", value="06", max_chars=2)
    with geo_cols[2]:
        county_fips = st.text_input("County FIPS", value="013", max_chars=3)
    with geo_cols[3]:
        include_proxy_fields = st.checkbox("include proxy fields", value=False)

    selectable_census_fields = supported_census_candidate_fields(include_proxy_fields=include_proxy_fields)
    selected_census_fields = st.multiselect(
        "Census fields",
        options=selectable_census_fields,
        default=[field for field in census_default_fields if field in selectable_census_fields],
    )
    if st.button("Fetch Census candidates"):
        try:
            census_candidates = fetch_census_source_candidates(
                state_fips=state_fips,
                county_fips=county_fips,
                year=int(census_year),
                fields=selected_census_fields,
                include_proxy_fields=include_proxy_fields,
            )
            st.session_state["api_source_candidates"]["census"] = census_candidates
            st.session_state["api_source_reports"]["census"] = census_candidates
            st.success(f"Fetched {len(census_candidates)} Census candidate fields")
        except CensusApiError as exc:
            st.error("Could not fetch Census ACS data.")
            st.exception(exc)

bea_default_fields = [field for field in required_field_names if field in supported_bea_candidate_fields()]
with st.expander("BEA Regional", expanded=False):
    bea_geo_cols = st.columns(4)
    with bea_geo_cols[0]:
        bea_year = st.number_input("BEA year", min_value=2001, max_value=2026, value=2024, step=1)
    with bea_geo_cols[1]:
        bea_prior_year = st.number_input("BEA prior year", min_value=2000, max_value=2025, value=2023, step=1)
    with bea_geo_cols[2]:
        bea_state_fips = st.text_input("BEA State FIPS", value=state_fips or "06", max_chars=2)
    with bea_geo_cols[3]:
        bea_county_fips = st.text_input("BEA County FIPS", value=county_fips or "013", max_chars=3)

    selected_bea_fields = st.multiselect(
        "BEA fields",
        options=supported_bea_candidate_fields(),
        default=bea_default_fields,
    )
    if st.button("Fetch BEA candidates"):
        try:
            bea_candidates = fetch_bea_source_candidates(
                state_fips=bea_state_fips,
                county_fips=bea_county_fips,
                year=int(bea_year),
                prior_year=int(bea_prior_year),
                fields=selected_bea_fields,
            )
            st.session_state["api_source_candidates"]["bea"] = bea_candidates
            st.session_state["api_source_reports"]["bea"] = bea_candidates
            st.success(f"Fetched {len(bea_candidates)} BEA candidate fields")
        except BeaApiError as exc:
            st.error("Could not fetch BEA Regional data.")
            st.exception(exc)

api_candidate_frames = [
    frame
    for frame in st.session_state.get("api_source_candidates", {}).values()
    if isinstance(frame, pd.DataFrame) and not frame.empty
]
if api_candidate_frames:
    st.dataframe(_display_df(pd.concat(api_candidate_frames, ignore_index=True)), width="stretch", hide_index=True)

st.divider()
st.subheader("Manual Fallback Editor")
st.caption(
    "This is optional. It lists formula raw fields so you can type values only when no upload/API source supplies them. "
    "Blank rows are ignored and do not become calculator inputs."
)

prefill_manual = st.checkbox(
    "Pre-fill manual values from current issuer_data",
    value=False,
    help="Use only when you intentionally want the currently saved issuer_data to become editable manual input.",
)
existing = (st.session_state.get("issuer_data", {}) or {}) if prefill_manual else {}
rows = []
for _, row in required_fields.iterrows():
    field_name = str(row["field_name"])
    rows.append(
        {
            "field_name": field_name,
            "value": existing.get(field_name, ""),
            "used_by": row.get("used_by", ""),
            "category": row.get("category", ""),
        }
    )

manual_df = pd.DataFrame(rows)
if not manual_df.empty:
    manual_df["value"] = manual_df["value"].apply(lambda x: "" if x is None else str(x))
if manual_df.empty:
    st.info("No formula-driven raw fields found for the selected methodology.")
    edited = manual_df
else:
    st.info(
        "For the true missing list, save issuer_data first and review Source Readiness Detail below. "
        "The calculator only reports fields missing from the selected final issuer_data, not every blank row in this editor."
    )
    edited = st.data_editor(
        manual_df,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"manual_source_editor_{methodology_id}",
        column_config={
            "value": st.column_config.TextColumn("value"),
            "used_by": st.column_config.TextColumn("used_by", disabled=True),
            "category": st.column_config.TextColumn("category", disabled=True),
        },
    )

if st.button("Save issuer_data", type="primary"):
    manual_data = _clean_edited_values(edited)
    manual_candidates = manual_data_to_source_candidates(manual_data)
    sourcing_output = run_data_sourcing_pipeline(
        [*mapped_candidate_frames, *api_candidate_frames, manual_candidates],
        methodology_id=methodology_id,
        required_fields=required_field_names,
        source_priority_path="config/source_priority.csv",
    )
    st.session_state["issuer_data"] = sourcing_output["issuer_data"]
    st.session_state["source_report"] = sourcing_output["source_report"]
    st.session_state["source_candidates"] = sourcing_output["source_candidates"]
    st.session_state["source_readiness_summary"] = sourcing_output["source_readiness_summary"]
    st.session_state["source_match_reports"] = pd.concat(match_reports, ignore_index=True) if match_reports else pd.DataFrame()
    st.success(f"Saved {len(sourcing_output['issuer_data'])} canonical fields. Go to Calculators next.")
    readiness_action(sourcing_output["source_report"])

with st.expander("Source Readiness Detail", expanded=True):
    readiness = st.session_state.get("source_readiness_summary", pd.DataFrame())
    source_report = st.session_state.get("source_report", pd.DataFrame())
    if isinstance(readiness, pd.DataFrame) and not readiness.empty:
        st.dataframe(_display_df(readiness), width="stretch", hide_index=True)
    else:
        st.info("No source selection has been saved yet.")
    if isinstance(source_report, pd.DataFrame) and not source_report.empty:
        selected_report = selected_source_report(source_report)
        tab1, tab2, tab3, tab4 = st.tabs(["Missing", "Ready", "Review", "All Selected"])
        with tab1:
            missing = selected_report[selected_report["readiness_status"].astype(str).eq("missing")]
            st.dataframe(_display_df(missing), width="stretch", hide_index=True) if not missing.empty else st.info("No selected fields are missing.")
        with tab2:
            ready = selected_report[selected_report["readiness_status"].astype(str).eq("independent_ready")]
            st.dataframe(_display_df(ready), width="stretch", hide_index=True) if not ready.empty else st.info("No independently ready fields yet.")
        with tab3:
            review = selected_report[
                ~selected_report["readiness_status"].astype(str).isin(["missing", "independent_ready"])
            ]
            st.dataframe(_display_df(review), width="stretch", hide_index=True) if not review.empty else st.info("No source-pending or review fields.")
        with tab4:
            st.dataframe(_display_df(selected_report), width="stretch", hide_index=True)

with st.expander("Current issuer_data", expanded=False):
    data = st.session_state.get("issuer_data", {}) or {}
    if data:
        st.json(data)
    else:
        st.info("No issuer_data saved yet.")
