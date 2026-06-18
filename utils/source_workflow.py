from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from connectors.bea_api import BeaApiError, fetch_bea_source_candidates, get_bea_api_key, supported_bea_candidate_fields
from connectors.census_api import (
    CensusApiError,
    fetch_census_source_candidates,
    get_census_api_key,
    supported_census_candidate_fields,
)
from connectors.creditscope_loader import load_creditscope_source_candidates
from engine.calculator_engine import calculate_all_formulas, load_formula_library, parse_required_fields
from engine.data_sourcing_engine import (
    DIRECT_METRIC_SOURCE_FIELDS,
    mapping_report_to_source_candidates,
    manual_data_to_source_candidates,
    required_fields_for_methodology,
    run_data_sourcing_pipeline,
)
from engine.factor_engine import load_factor_template
from engine.mapping_engine import map_uploaded_file
from utils.ui_helpers import clean_for_display, selected_source_report, source_readiness_counts


SOURCE_SESSION_KEYS = {
    "issuer_data",
    "source_report",
    "source_candidates",
    "source_readiness_summary",
    "source_match_reports",
    "uploaded_source_candidates",
    "uploaded_source_reports",
    "uploaded_sources",
    "uploaded_pdf_documents",
    "acfr_pdf_pages_cache",
    "last_acfr_auto_snippets",
    "uploaded_issuer_data",
    "api_source_candidates",
    "api_source_reports",
    "manual_source_candidates",
    "manual_source_values",
    "approved_source_candidates",
    "workbook_direct_metric_debug",
}

SOURCE_WORKFLOW_CACHE_VERSION = "issuer-input-raw-fields-v1"
LATEST_CENSUS_SOURCE_YEAR = 2024
LATEST_BEA_SOURCE_YEAR = 2024
ISSUER_DATA_EDITOR_EXCLUDED_FIELDS = set(DIRECT_METRIC_SOURCE_FIELDS)


SOURCE_WORKFLOW_GUIDE: list[dict[str, str]] = [
    {
        "section": "Source uploads",
        "what_it_does": "Registers files and extracts source candidates when the file is mappable. CreditScope workbook values feed scoring after Save issuer_data.",
        "when_to_use": "Start here for CreditScope workbook, ACFR, official statement/debt support, or IPEDS files.",
        "what_it_does_not_do": "ACFR/OS PDFs are evidence files. They do not automatically replace formula inputs until evidence is approved and applied.",
    },
    {
        "section": "API candidates",
        "what_it_does": "Fetches Census and BEA candidate values for economy, population, income, and demographic fields.",
        "when_to_use": "Use after deal setup when geography/year are known.",
        "what_it_does_not_do": "Fetched API rows still need Save issuer_data before formulas can use them.",
    },
    {
        "section": "Manual / source-pending inputs",
        "what_it_does": "Lets the analyst type missing raw values and then saves the selected source candidates into issuer_data.",
        "when_to_use": "Use after uploads/API fetches, especially when Blocking Required fields remain missing.",
        "what_it_does_not_do": "Blank cells do not overwrite uploaded/API values.",
    },
    {
        "section": "Source inventory readiness",
        "what_it_does": "Shows extraction coverage for selected source rows.",
        "when_to_use": "Use it to understand source inventory quality before Data Confirmation.",
        "what_it_does_not_do": "This is not the final rating blocker list; Data Confirmation decides what actually blocks scoring.",
    },
]


def _uploaded_file_payload(uploaded_file: Any) -> tuple[str, bytes]:
    name = str(getattr(uploaded_file, "name", "") or "uploaded_file")
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    if hasattr(uploaded_file, "getvalue"):
        payload = uploaded_file.getvalue()
    else:
        payload = uploaded_file.read()
    if isinstance(payload, str):
        payload = payload.encode("utf-8")
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    return name, bytes(payload)


def _upload_buffer(file_name: str, payload: bytes) -> io.BytesIO:
    buffer = io.BytesIO(payload)
    buffer.name = file_name
    return buffer


def _show_api_fetch_error(source_label: str, exc: Exception) -> None:
    message = str(exc).strip() or "No detail returned."
    st.warning(f"{source_label} candidate data is unavailable for the selected inputs.")
    st.caption(f"{type(exc).__name__}: {message}")
    if source_label == "Census ACS":
        st.caption("Try a published ACS year, confirm state/county FIPS, or use manual/source confirmation for this run.")
    elif source_label == "BEA":
        st.caption("Confirm the BEA API key, year, and county FIPS. You can continue without BEA candidates and fill economy fields later.")


def _show_missing_api_key(source_label: str, secret_name: str) -> None:
    st.info(f"{source_label} is optional and is currently disabled because `{secret_name}` is not configured.")
    st.caption(f"To enable it on Streamlit Cloud, add `{secret_name}` in App settings > Secrets, then reboot the app.")


@st.cache_data(show_spinner=False)
def _cached_required_names(methodology_id: str) -> tuple[str, ...]:
    return tuple(required_fields_for_methodology(methodology_id))


@st.cache_data(show_spinner=False)
def _cached_excel_sheet_names(file_name: str, payload: bytes) -> List[str]:
    if Path(file_name).suffix.lower() not in {".xlsx", ".xls"}:
        return []
    workbook = None
    try:
        workbook = load_workbook(_upload_buffer(file_name, payload), read_only=True, data_only=True)
        return list(workbook.sheetnames)
    except Exception:
        return []
    finally:
        if workbook is not None:
            workbook.close()


@st.cache_data(show_spinner=False)
def _cached_creditscope_mapping(
    file_name: str,
    payload: bytes,
    sheet_name: str | None,
    required_fields: tuple[str, ...],
    include_support_tabs: bool,
    cache_version: str,
) -> Dict[str, Any]:
    return load_creditscope_source_candidates(
        uploaded_file=_upload_buffer(file_name, payload),
        mapping_path="config/field_mapping.csv",
        row_mapping_path="config/creditscope_row_mapping.csv",
        sheet_name=sheet_name,
        value_col=2,
        required_fields=list(required_fields),
        include_support_tabs=include_support_tabs,
    )


@st.cache_data(show_spinner=False)
def _cached_mapped_upload(
    file_name: str,
    payload: bytes,
    source_name: str,
) -> tuple[Dict[str, Any], pd.DataFrame, pd.DataFrame]:
    source_data, report = map_uploaded_file(
        uploaded_file=_upload_buffer(file_name, payload),
        source_name=source_name,
        mapping_path="config/field_mapping.csv",
    )
    candidates = mapping_report_to_source_candidates(report, uploaded_file=file_name)
    return source_data, report, candidates


@st.cache_data(ttl=86400, show_spinner=False)
def _cached_census_source_candidates(
    state_fips: str,
    county_fips: str,
    year: int,
    fields: tuple[str, ...],
    include_proxy_fields: bool,
) -> pd.DataFrame:
    return fetch_census_source_candidates(
        state_fips=state_fips,
        county_fips=county_fips,
        year=year,
        fields=list(fields),
        include_proxy_fields=include_proxy_fields,
    )


@st.cache_data(ttl=86400, show_spinner=False)
def _cached_bea_source_candidates(
    state_fips: str,
    county_fips: str,
    year: int,
    prior_year: int,
    fields: tuple[str, ...],
) -> pd.DataFrame:
    return fetch_bea_source_candidates(
        state_fips=state_fips,
        county_fips=county_fips,
        year=year,
        prior_year=prior_year,
        fields=list(fields),
    )


def _reset_source_session(methodology_id: str) -> None:
    for key in SOURCE_SESSION_KEYS:
        st.session_state.pop(key, None)
    for key in list(st.session_state.keys()):
        if str(key).startswith(("upload_", "sheet_", "manual_source_editor_", "api_")):
            st.session_state.pop(key, None)
    st.session_state["uploaded_sources"] = {}
    st.session_state["uploaded_pdf_documents"] = {}
    st.session_state["uploaded_source_candidates"] = {}
    st.session_state["uploaded_source_reports"] = {}
    st.session_state["uploaded_issuer_data"] = {}
    st.session_state["api_source_candidates"] = {}
    st.session_state["api_source_reports"] = {}
    st.session_state["manual_source_values"] = {}
    st.session_state["approved_source_candidates"] = pd.DataFrame()
    st.session_state["source_reset_notice"] = "Source session reset. Upload/fetch sources again before saving issuer_data."
    st.session_state["source_methodology_id"] = methodology_id


def _excel_sheet_names(uploaded_file: Any) -> List[str]:
    name, payload = _uploaded_file_payload(uploaded_file)
    return _cached_excel_sheet_names(name, payload)


def _tokens(value: str) -> set[str]:
    ignored = {
        "scorecard",
        "moodys",
        "moody",
        "higher",
        "local",
        "gov",
        "water",
        "sewer",
        "utility",
        "xlsx",
        "xls",
        "raw",
        "credit",
        "scope",
        "creditscope",
        "2023",
        "2024",
        "2025",
        "2026",
        "fin",
        "ccd",
        "go",
        "sp",
        "k12",
    }
    return {
        token
        for token in re.findall(r"[a-z0-9]+", str(value).lower())
        if len(token) > 2 and token not in ignored
    }


def _raw_hint_score(sheet_name: str) -> int:
    lowered = str(sheet_name).lower()
    score = 0
    if any(token in lowered for token in ["fin", "raw", "creditscope", "credit scope"]):
        score += 2
    if any(token in lowered for token in ["scorecard", "public", "summary", "validation"]):
        score -= 3
    return score


def _is_generic_raw_sheet(sheet_name: str) -> bool:
    lowered = re.sub(r"\s+", " ", str(sheet_name).strip().lower())
    return lowered in {"creditscope", "credit scope", "raw", "fin", "issuer data"}


def _auto_sheet(sheet_names: Iterable[str], uploaded_name: str) -> str | None:
    sheet_names = list(sheet_names)
    if len(sheet_names) == 1:
        return sheet_names[0]
    upload_tokens = _tokens(uploaded_name)
    ranked: list[tuple[int, int, int, str]] = []
    for idx, sheet in enumerate(sheet_names):
        sheet_tokens = _tokens(sheet)
        overlap = len(upload_tokens & sheet_tokens)
        hint = _raw_hint_score(sheet)
        if overlap > 0 and hint >= 0:
            ranked.append((overlap, hint, -idx, sheet))
    if ranked:
        return max(ranked)[3]
    generic_exact = [sheet for sheet in sheet_names if _is_generic_raw_sheet(sheet)]
    if len(generic_exact) == 1:
        return generic_exact[0]
    generic = [sheet for sheet in sheet_names if _raw_hint_score(sheet) > 0 and not _tokens(sheet)]
    return generic[0] if len(generic) == 1 else None


def _build_required_field_frame(methodology_id: str) -> pd.DataFrame:
    formulas = load_formula_library("config/formula_library.csv")
    template = load_factor_template(methodology_id, templates_dir="templates")
    formula_ids = set(template["formula_id"].dropna().astype(str))
    rows: list[dict[str, Any]] = []
    for _, formula in formulas[formulas["formula_id"].astype(str).isin(formula_ids)].iterrows():
        for field in parse_required_fields(formula.get("required_data", "")):
            if field == "manual":
                continue
            rows.append(
                {
                    "field_name": field,
                    "used_by": str(formula.get("formula_id", "")),
                    "category": str(formula.get("category", "")),
                }
            )
    if not rows:
        return pd.DataFrame(columns=["field_name", "used_by", "category"])
    return (
        pd.DataFrame(rows)
        .groupby("field_name", as_index=False)
        .agg(
            used_by=("used_by", lambda x: "; ".join(sorted(set(v for v in x if v)))),
            category=("category", lambda x: "; ".join(sorted(set(v for v in x if v)))),
        )
        .sort_values("field_name")
        .reset_index(drop=True)
    )


@st.cache_data(show_spinner=False)
def _required_field_frame(methodology_id: str) -> pd.DataFrame:
    return _build_required_field_frame(methodology_id)


@st.cache_data(show_spinner=False)
def _cached_methodology_formula_ids(methodology_id: str) -> tuple[str, ...]:
    template = load_factor_template(methodology_id, templates_dir="templates")
    return tuple(sorted(set(template["formula_id"].dropna().astype(str).str.strip()) - {""}))


def _has_source_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    try:
        return not bool(pd.isna(value))
    except (TypeError, ValueError):
        return True


def _workbook_direct_metric_overrides(methodology_id: str) -> Dict[str, Dict[str, Any]]:
    formula_ids = set(_cached_methodology_formula_ids(methodology_id))
    uploads = st.session_state.get("uploaded_issuer_data", {}) or {}
    candidate_uploads = st.session_state.get("uploaded_source_candidates", {}) or {}
    uploaded_sources = st.session_state.get("uploaded_sources", {}) or {}
    overrides: Dict[str, Dict[str, Any]] = {}
    for upload_key, payload in uploads.items():
        if not isinstance(payload, dict):
            continue
        if str(payload.get("source_name", "")).strip() != "CreditScope":
            continue
        workbook_data = payload.get("issuer_data", {}) or {}
        if not isinstance(workbook_data, dict):
            continue
        source_name = str(payload.get("source_name") or upload_key)
        file_name = str(payload.get("file_name") or "").strip()
        source_used = f"{source_name}: {file_name}" if file_name else source_name
        for field, value in workbook_data.items():
            field_name = str(field).strip()
            if field_name not in formula_ids or not _has_source_value(value):
                continue
            overrides[field_name] = {
                "source_used": source_used,
                "workbook_value": value,
            }
        candidates = candidate_uploads.get(upload_key)
        if not isinstance(candidates, pd.DataFrame) or candidates.empty:
            continue
        for _, row in candidates.iterrows():
            field_name = str(row.get("field_name", "") or "").strip()
            value = row.get("value")
            if field_name not in formula_ids or not _has_source_value(value):
                continue
            row_source = str(row.get("source_name") or row.get("canonical_source") or source_used).strip()
            row_detail = str(row.get("source_cell_or_api") or row.get("source_detail") or "").strip()
            overrides[field_name] = {
                "source_used": f"{row_source}: {row_detail}" if row_detail else row_source,
                "workbook_value": value,
            }
    for upload_key, candidates in candidate_uploads.items():
        if upload_key in uploads:
            continue
        if str(upload_key).strip().lower() != "creditscope":
            continue
        if not isinstance(candidates, pd.DataFrame) or candidates.empty:
            continue
        file_name = str(uploaded_sources.get(upload_key, "") or "").strip()
        fallback_source = f"CreditScope: {file_name}" if file_name else "CreditScope"
        for _, row in candidates.iterrows():
            field_name = str(row.get("field_name", "") or "").strip()
            value = row.get("value")
            if field_name not in formula_ids or not _has_source_value(value):
                continue
            row_source = str(row.get("source_name") or row.get("canonical_source") or fallback_source).strip()
            row_detail = str(row.get("source_cell_or_api") or row.get("source_detail") or "").strip()
            overrides[field_name] = {
                "source_used": f"{row_source}: {row_detail}" if row_detail else row_source,
                "workbook_value": value,
            }
    return overrides


def _direct_metric_debug_frame(
    overrides: Dict[str, Dict[str, Any]],
    issuer_data: Dict[str, Any],
) -> pd.DataFrame:
    rows = [
        {
            "field_name": field,
            "source_used": item.get("source_used", "workbook_direct_metric"),
            "workbook_value": item.get("workbook_value"),
            "final_formula_input": issuer_data.get(field),
        }
        for field, item in sorted(overrides.items())
    ]
    return pd.DataFrame(rows, columns=["field_name", "source_used", "workbook_value", "final_formula_input"])


def _clear_formula_rating_outputs() -> None:
    for key in ["formula_results", "methodology_formula_results", "rating_output"]:
        st.session_state.pop(key, None)
    st.session_state["formula_results"] = pd.DataFrame()
    st.session_state["methodology_formula_results"] = pd.DataFrame()
    st.session_state["rating_output"] = None


def _complete_required_field_frame(required_fields: pd.DataFrame, required_names: Iterable[str]) -> pd.DataFrame:
    existing = set(required_fields["field_name"].dropna().astype(str)) if not required_fields.empty else set()
    missing = sorted({str(field) for field in required_names if str(field)} - existing)
    if not missing:
        return required_fields
    additions = pd.DataFrame(
        [{"field_name": field, "used_by": "", "category": ""} for field in missing]
    )
    return (
        pd.concat([required_fields, additions], ignore_index=True)
        .drop_duplicates(subset=["field_name"], keep="first")
        .sort_values("field_name")
        .reset_index(drop=True)
    )


def _parse_year(value: Any, fallback: int) -> int:
    match = re.search(r"\d{4}", str(value or ""))
    if not match:
        return fallback
    try:
        return int(match.group(0))
    except Exception:
        return fallback


def _bounded_year(value: Any, *, minimum: int, maximum: int, fallback: int) -> int:
    parsed = _parse_year(value, fallback)
    return max(minimum, min(maximum, parsed))


def _source_year_defaults() -> tuple[str, int, int, int]:
    analysis_year = str(st.session_state.get("analysis_year", "") or "").strip()
    parsed_analysis_year = _parse_year(analysis_year, LATEST_CENSUS_SOURCE_YEAR)
    census_year = _bounded_year(
        parsed_analysis_year,
        minimum=2009,
        maximum=LATEST_CENSUS_SOURCE_YEAR,
        fallback=LATEST_CENSUS_SOURCE_YEAR,
    )
    bea_year = _bounded_year(
        parsed_analysis_year,
        minimum=2001,
        maximum=LATEST_BEA_SOURCE_YEAR,
        fallback=LATEST_BEA_SOURCE_YEAR,
    )
    bea_prior = max(2000, min(bea_year - 1, LATEST_BEA_SOURCE_YEAR - 1))
    return analysis_year or str(parsed_analysis_year), census_year, bea_year, bea_prior


def _recommended_api_defaults() -> dict[str, Any]:
    analysis_year_label, census_year, bea_year, bea_prior = _source_year_defaults()
    issuer_name = str(st.session_state.get("issuer_name", "") or "").lower()
    default_state_fips = str(st.session_state.get("state_fips", "") or "06").zfill(2)
    stored_county_fips = str(st.session_state.get("county_fips", "") or "").strip()
    is_west_sacramento = "west sacramento" in issuer_name
    if is_west_sacramento and stored_county_fips in {"", "013"}:
        default_county_fips = "113"
    else:
        default_county_fips = str(stored_county_fips or "013").zfill(3)
    return {
        "analysis_year_label": analysis_year_label,
        "census_year": census_year,
        "bea_year": bea_year,
        "bea_prior": bea_prior,
        "state_fips": default_state_fips,
        "county_fips": default_county_fips,
        "is_west_sacramento": is_west_sacramento,
    }


def _set_recommended_api_widget_defaults(defaults: dict[str, Any], *, force: bool = False) -> None:
    simple_defaults = {
        "api_census_source_year": defaults["census_year"],
        "api_bea_source_year": defaults["bea_year"],
        "api_state_fips": defaults["state_fips"],
        "api_county_fips": defaults["county_fips"],
    }
    for key, value in simple_defaults.items():
        if force or key not in st.session_state:
            st.session_state[key] = value

    if defaults.get("is_west_sacramento"):
        for key in ["api_county_fips"]:
            current = str(st.session_state.get(key, "") or "").zfill(3)
            if force or current in {"", "013"}:
                st.session_state[key] = "113"

    current_bea_year = _bounded_year(
        st.session_state.get("api_bea_source_year"),
        minimum=2001,
        maximum=LATEST_BEA_SOURCE_YEAR,
        fallback=int(defaults["bea_year"]),
    )
    if force or current_bea_year == int(defaults["bea_prior"]):
        st.session_state["api_bea_source_year"] = int(defaults["bea_year"])


def _source_candidate_frames(include_manual_values: bool) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    frames.extend(
        frame
        for frame in st.session_state.get("uploaded_source_candidates", {}).values()
        if isinstance(frame, pd.DataFrame) and not frame.empty
    )
    frames.extend(
        frame
        for frame in st.session_state.get("api_source_candidates", {}).values()
        if isinstance(frame, pd.DataFrame) and not frame.empty
    )
    approved_candidates = st.session_state.get("approved_source_candidates")
    if isinstance(approved_candidates, pd.DataFrame) and not approved_candidates.empty:
        frames.append(approved_candidates)
    if include_manual_values:
        manual_values = st.session_state.get("manual_source_values", {}) or {}
        manual_candidates = manual_data_to_source_candidates(manual_values)
        if not manual_candidates.empty:
            frames.append(manual_candidates)
        st.session_state["manual_source_candidates"] = manual_candidates
    return frames


def _clean_input_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if str(value).strip() == "":
        return None
    try:
        return float(value)
    except Exception:
        return str(value).strip()


def _editor_text_value(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value)


def _values_match(left: Any, right: Any) -> bool:
    left_clean = _clean_input_value(left)
    right_clean = _clean_input_value(right)
    if left_clean is None and right_clean is None:
        return True
    if left_clean is None or right_clean is None:
        return False
    try:
        return abs(float(left_clean) - float(right_clean)) <= 1e-9
    except Exception:
        return str(left_clean).strip() == str(right_clean).strip()


def _source_label_from_row(row: pd.Series) -> str:
    parts = [
        str(row.get("canonical_source") or row.get("source_name") or "").strip(),
        str(row.get("source_file") or "").strip(),
        str(row.get("source_cell_or_api") or row.get("source_table") or "").strip(),
    ]
    return ": ".join(part for part in parts if part)


def _build_issuer_data_editor(
    *,
    methodology_id: str,
    required_fields: pd.DataFrame,
    required_names: list[str],
    frames: list[pd.DataFrame],
) -> tuple[pd.DataFrame, Dict[str, Any]]:
    result = run_data_sourcing_pipeline(
        frames,
        methodology_id=methodology_id,
        required_fields=required_names,
    )
    issuer_data = dict(result["issuer_data"])
    source_report = result["source_report"]
    selected = selected_source_report(source_report)
    selected_lookup: dict[str, pd.Series] = {}
    if isinstance(selected, pd.DataFrame) and not selected.empty and "field_name" in selected.columns:
        for _, row in selected.iterrows():
            field = str(row.get("field_name", "") or "").strip()
            if field and field not in selected_lookup:
                selected_lookup[field] = row

    direct_metric_overrides = _workbook_direct_metric_overrides(methodology_id)
    for field, item in direct_metric_overrides.items():
        issuer_data[field] = item.get("workbook_value")

    manual_values = st.session_state.get("manual_source_values", {}) or {}
    source_value_map: dict[str, Any] = {}
    rows: list[dict[str, Any]] = []
    for _, row in required_fields.iterrows():
        field = str(row.get("field_name", "") or "").strip()
        if not field:
            continue
        if field in ISSUER_DATA_EDITOR_EXCLUDED_FIELDS:
            continue
        selected_row = selected_lookup.get(field)
        status = "missing"
        source_used = ""
        value = issuer_data.get(field)
        if field in direct_metric_overrides:
            value = direct_metric_overrides[field].get("workbook_value")
            status = "workbook_direct_metric"
            source_used = str(direct_metric_overrides[field].get("source_used") or "CreditScope workbook").strip()
        elif value is not None:
            status = str(selected_row.get("readiness_status", "independent_ready") if selected_row is not None else "independent_ready")
            source_used = _source_label_from_row(selected_row) if selected_row is not None else ""
        elif field in manual_values and str(manual_values[field]).strip() != "":
            value = manual_values[field]
            status = "manual_input"
            source_used = "Manual"
        elif selected_row is not None:
            status = str(selected_row.get("readiness_status", "") or "missing")
            source_used = _source_label_from_row(selected_row)

        source_value_map[field] = value
        rows.append(
            {
                "field_name": field,
                "value": _editor_text_value(value),
                "source_status": status,
                "source_used": source_used,
                "used_by": row.get("used_by", ""),
                "category": row.get("category", ""),
            }
        )
    editor_df = pd.DataFrame(rows, columns=["field_name", "value", "source_status", "source_used", "used_by", "category"])
    return editor_df, {
        "source_result": result,
        "source_value_map": source_value_map,
        "direct_metric_debug": _direct_metric_debug_frame(direct_metric_overrides, issuer_data),
    }


def _save_issuer_data_from_editor(
    edited_inputs: pd.DataFrame,
    *,
    methodology_id: str,
    source_context: Dict[str, Any],
    run_formulas: bool,
) -> None:
    issuer_data: Dict[str, Any] = {}
    manual_values: Dict[str, Any] = {}
    source_value_map = source_context.get("source_value_map", {}) or {}
    if isinstance(edited_inputs, pd.DataFrame) and not edited_inputs.empty:
        for _, row in edited_inputs.iterrows():
            field = str(row.get("field_name", "") or "").strip()
            value = _clean_input_value(row.get("value"))
            if not field or value is None:
                continue
            issuer_data[field] = value
            if not _values_match(value, source_value_map.get(field)):
                manual_values[field] = value

    result = source_context.get("source_result", {}) or {}
    st.session_state["issuer_data"] = issuer_data
    st.session_state["manual_source_values"] = manual_values
    st.session_state["source_report"] = result.get("source_report", pd.DataFrame())
    st.session_state["source_candidates"] = result.get("source_candidates", pd.DataFrame())
    st.session_state["source_readiness_summary"] = result.get("source_readiness_summary", pd.DataFrame())
    st.session_state["workbook_direct_metric_debug"] = source_context.get("direct_metric_debug", pd.DataFrame())
    st.session_state["issuer_data_input_table"] = edited_inputs.copy() if isinstance(edited_inputs, pd.DataFrame) else pd.DataFrame()
    _clear_formula_rating_outputs()
    st.session_state["source_saved_needs_formula_run"] = True
    reports = []
    reports.extend(st.session_state.get("uploaded_source_reports", {}).values())
    reports.extend(st.session_state.get("api_source_reports", {}).values())
    if reports:
        st.session_state["source_match_reports"] = pd.concat(
            [r for r in reports if isinstance(r, pd.DataFrame) and not r.empty],
            ignore_index=True,
        )
    st.success(f"Saved issuer_data with {len(issuer_data)} calculation input fields.")
    manual_count = len(manual_values)
    if manual_count:
        st.caption(f"{manual_count} value(s) were saved as manual replacements because they were blank or changed from the suggested source value.")
    if run_formulas:
        formula_results = _run_formulas_after_source_save(methodology_id, issuer_data)
        st.success(f"Ran formulas from the saved issuer_data. {len(formula_results)} formula rows saved.")
    else:
        st.info("Formula and scoreboard outputs were cleared. Run formulas next when you are ready.")


def _save_manual_raw_gap_values(
    edited_inputs: pd.DataFrame,
    *,
    methodology_id: str,
    run_formulas: bool,
) -> None:
    manual_values = dict(st.session_state.get("manual_source_values", {}) or {})
    updated: Dict[str, Any] = {}
    if isinstance(edited_inputs, pd.DataFrame) and not edited_inputs.empty:
        for _, row in edited_inputs.iterrows():
            field = str(row.get("field_name", "") or "").strip()
            value = _clean_input_value(row.get("manual_value"))
            if not field or value is None:
                continue
            manual_values[field] = value
            updated[field] = value

    if not updated:
        st.warning("No manual values were entered.")
        return

    st.session_state["manual_source_values"] = manual_values
    frames = _source_candidate_frames(include_manual_values=True)
    required_names = required_fields_for_methodology(methodology_id)
    result = run_data_sourcing_pipeline(
        frames,
        methodology_id=methodology_id,
        required_fields=required_names,
    )
    issuer_data = result.get("issuer_data", {})
    st.session_state["issuer_data"] = issuer_data
    st.session_state["source_candidates"] = result.get("source_candidates", pd.DataFrame())
    st.session_state["source_report"] = result.get("source_report", pd.DataFrame())
    st.session_state["source_readiness_summary"] = result.get("source_readiness_summary", pd.DataFrame())
    _clear_formula_rating_outputs()
    st.session_state["source_saved_needs_formula_run"] = True
    st.session_state["manual_raw_gap_save_notice"] = f"Saved {len(updated)} manual raw value(s) into issuer_data."
    if run_formulas:
        formula_results = _run_formulas_after_source_save(methodology_id, issuer_data)
        st.session_state["manual_raw_gap_save_notice"] += f" Ran formulas from updated issuer_data; {len(formula_results)} formula rows saved."
    else:
        st.session_state["manual_raw_gap_save_notice"] += " Formula and scoreboard outputs were cleared."


def _manual_raw_gap_input_frame(missing: pd.DataFrame) -> pd.DataFrame:
    manual_values = st.session_state.get("manual_source_values", {}) or {}
    rows: list[dict[str, Any]] = []
    for _, row in missing.iterrows():
        field = str(row.get("field_name", "") or "").strip()
        if not field:
            continue
        manual_value = _clean_input_value(manual_values.get(field))
        rows.append(
            {
                "field_name": field,
                "manual_value": manual_value,
                "unit": row.get("unit", "") or "dollars",
                "notes": (
                    "For West Sacramento, use 0 only if confirming the scorecard has no transfer adjustment."
                    if field == "operating_transfers"
                    else ""
                ),
            }
        )
    out = pd.DataFrame(rows, columns=["field_name", "manual_value", "unit", "notes"])
    if "manual_value" in out.columns:
        out["manual_value"] = pd.to_numeric(out["manual_value"], errors="coerce")
    return out


def _render_manual_raw_gap_form(missing: pd.DataFrame, methodology_id: str) -> None:
    if not isinstance(missing, pd.DataFrame) or missing.empty:
        return
    with st.form(f"manual_raw_gap_form_{methodology_id}"):
        st.markdown("**Manual raw value entry**")
        st.caption(
            "Use this only when the analyst has confirmed the value from ACFR/OS/CreditScope support or a policy assumption. "
            "Saved manual values become Manual source candidates and flow into issuer_data. Scalar manual entries resolve "
            "source gaps; exact avg_3yr replication may still require three-year source series."
        )
        manual_inputs = st.data_editor(
            _manual_raw_gap_input_frame(missing),
            width="stretch",
            hide_index=True,
            num_rows="fixed",
            key=f"manual_raw_gap_editor_{methodology_id}",
            column_config={
                "field_name": st.column_config.TextColumn("field_name", disabled=True),
                "manual_value": st.column_config.NumberColumn("manual_value"),
                "unit": st.column_config.TextColumn("unit", disabled=True),
                "notes": st.column_config.TextColumn("notes", disabled=True),
            },
        )
        cols = st.columns(2)
        save_manual = cols[0].form_submit_button("Save manual raw values", type="primary")
        save_manual_run = cols[1].form_submit_button("Save manual values and run formulas")

    if save_manual or save_manual_run:
        _save_manual_raw_gap_values(
            manual_inputs,
            methodology_id=methodology_id,
            run_formulas=save_manual_run,
        )
        st.rerun()


def _readiness_tabs(source_report: pd.DataFrame, methodology_id: str) -> None:
    if not isinstance(source_report, pd.DataFrame) or source_report.empty:
        st.info("No source report saved yet.")
        return
    notice = st.session_state.pop("manual_raw_gap_save_notice", "")
    if notice:
        st.success(notice)
    selected = selected_source_report(source_report)
    if "field_name" in selected.columns:
        direct_metric_mask = selected["field_name"].astype(str).isin(ISSUER_DATA_EDITOR_EXCLUDED_FIELDS)
    else:
        direct_metric_mask = pd.Series(False, index=selected.index)
    raw_selected = selected[~direct_metric_mask].copy()
    calculated_metrics = selected[direct_metric_mask].copy()
    counts = source_readiness_counts(raw_selected)
    status_labels = {
        "missing": "support_missing (non-blocking unless also a formula blocker)",
        "independent_ready": "independent_ready",
        "source_pending": "support_pending",
        "needs_review": "support_review",
    }
    st.caption(
        "Support inventory only: this is pre-formula extraction coverage. Support Missing rows are raw evidence gaps, "
        "not rating blockers when a formula-ready value already feeds scoring. Calculated/direct metrics are shown "
        "separately because they are formula outputs or official support-tab checks, not manual source inputs."
    )
    summary_rows = [
        {
            "inventory_status": status_labels.get(key, key),
            "field_count": value,
        }
        for key, value in counts.items()
    ]
    if not calculated_metrics.empty:
        calculated_count = (
            calculated_metrics["field_name"].nunique()
            if "field_name" in calculated_metrics.columns
            else len(calculated_metrics)
        )
        summary_rows.append(
            {
                "inventory_status": "calculated_or_direct_metric",
                "field_count": calculated_count,
            }
        )
    st.dataframe(pd.DataFrame(summary_rows), width="stretch", hide_index=True)
    missing = raw_selected[raw_selected["readiness_status"].astype(str).eq("missing")]
    ready = raw_selected[raw_selected["readiness_status"].astype(str).eq("independent_ready")]
    review = raw_selected[raw_selected["readiness_status"].astype(str).isin(["source_pending", "needs_review"])]
    tabs = st.tabs(["Raw Missing", "Inventory Ready", "Inventory Review", "Calculated Metrics", "All Selected"])
    for idx, (tab, frame, empty) in enumerate([
        (tabs[0], missing, "No missing raw support fields."),
        (tabs[1], ready, "No independently ready fields yet."),
        (tabs[2], review, "No source-pending or review fields."),
        (tabs[3], calculated_metrics, "No calculated/direct metric rows."),
        (tabs[4], selected, "No selected source rows."),
    ]):
        with tab:
            if frame.empty:
                st.info(empty)
            else:
                if idx == 0:
                    st.info(
                        "These fields were not separately extracted from the raw source inventory. "
                        "They are the remaining source fields to confirm or fill before formula scoring is complete."
                    )
                elif idx == 3:
                    st.info(
                        "These rows are calculated later from raw inputs, or used only when validating an official "
                        "support-tab direct metric. They are not manual issuer_data inputs."
                    )
                st.dataframe(clean_for_display(frame), width="stretch", hide_index=True)
                if idx == 0:
                    _render_manual_raw_gap_form(frame, methodology_id)


def _uploaded_sources_summary() -> pd.DataFrame:
    uploads = st.session_state.get("uploaded_sources", {}) or {}
    rows: list[dict[str, Any]] = []
    for source_slot, file_names in uploads.items():
        if not str(file_names or "").strip():
            continue
        candidates = st.session_state.get("uploaded_source_candidates", {}).get(source_slot)
        reports = st.session_state.get("uploaded_source_reports", {}).get(source_slot)
        issuer_payload = st.session_state.get("uploaded_issuer_data", {}).get(source_slot, {}) or {}
        issuer_data = issuer_payload.get("issuer_data", {}) if isinstance(issuer_payload, dict) else {}
        rows.append(
            {
                "source_slot": source_slot,
                "saved_files": file_names,
                "mapped_fields": len(issuer_data) if isinstance(issuer_data, dict) else 0,
                "candidate_rows": len(candidates) if isinstance(candidates, pd.DataFrame) else 0,
                "diagnostic_rows": len(reports) if isinstance(reports, pd.DataFrame) else 0,
                "session_status": "saved in session",
            }
        )
    return pd.DataFrame(rows)


def _save_uploaded_pdf_document(source_slot: str, source_name: str, file_name: str, payload: bytes) -> None:
    store = st.session_state.setdefault("uploaded_pdf_documents", {})
    docs = list(store.get(source_slot, []))
    incoming_size = len(payload or b"")
    docs = [
        doc
        for doc in docs
        if not (
            str(doc.get("file_name", "")) == file_name
            and int(doc.get("file_size", -1) or -1) == incoming_size
        )
    ]
    docs.append(
        {
            "source_slot": source_slot,
            "source_name": source_name,
            "file_name": file_name,
            "file_size": incoming_size,
            "payload": payload,
        }
    )
    store[source_slot] = docs
    st.session_state["uploaded_pdf_documents"] = store


def _methodology_formula_results(methodology_id: str, formula_results: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(formula_results, pd.DataFrame) or formula_results.empty:
        return pd.DataFrame()
    try:
        template = load_factor_template(methodology_id, templates_dir="templates")
    except Exception:
        return formula_results.copy()
    if template.empty or "formula_id" not in template.columns or "formula_id" not in formula_results.columns:
        return formula_results.copy()
    ids = set(template["formula_id"].dropna().astype(str))
    return formula_results[formula_results["formula_id"].astype(str).isin(ids)].copy()


def _run_formulas_after_source_save(methodology_id: str, issuer_data: dict[str, Any]) -> pd.DataFrame:
    formula_results = calculate_all_formulas(issuer_data)
    st.session_state["formula_results"] = formula_results
    st.session_state["methodology_formula_results"] = _methodology_formula_results(methodology_id, formula_results)
    st.session_state["rating_output"] = None
    st.session_state["source_saved_needs_formula_run"] = False
    return formula_results


def _source_guided_progress() -> pd.DataFrame:
    uploads = st.session_state.get("uploaded_sources", {}) or {}
    issuer_data = st.session_state.get("issuer_data", {}) or {}
    approved = st.session_state.get("approved_source_candidates")
    approved_count = len(approved) if isinstance(approved, pd.DataFrame) else 0
    rows = [
        {
            "step": "1",
            "task": "Upload CreditScope financial workbook",
            "status": "Done" if uploads.get("creditscope") else "Next",
            "what_to_do": "Put Financials_ West Sacramento.xlsx in CreditScope raw workbook.",
        },
        {
            "step": "2",
            "task": "Upload ACFR and debt support PDFs",
            "status": "Done" if uploads.get("acfr") or uploads.get("debt_report") else "After CreditScope",
            "what_to_do": "Put ACFR PDFs in ACFR; put debt schedules in Debt service / bonded indebtedness report.",
        },
        {
            "step": "3",
            "task": "Save issuer_data",
            "status": "Done" if issuer_data else "After uploads",
            "what_to_do": "Click Save uploaded data as issuer_data. This does not approve PDF values yet.",
        },
        {
            "step": "4",
            "task": "Confirm pending source values",
            "status": "In progress" if approved_count else "Later",
            "what_to_do": "Open Source Confirmation Queue and accept/edit only values you trust.",
        },
    ]
    return pd.DataFrame(rows)


def render_source_workflow(methodology_id: str) -> None:
    if st.session_state.get("source_methodology_id") != methodology_id:
        st.session_state["api_source_candidates"] = {}
        st.session_state["api_source_reports"] = {}
        st.session_state["uploaded_source_candidates"] = {}
        st.session_state["uploaded_source_reports"] = {}
        st.session_state["approved_source_candidates"] = pd.DataFrame()
        st.session_state["source_methodology_id"] = methodology_id
    st.session_state.setdefault("uploaded_sources", {})
    st.session_state.setdefault("uploaded_pdf_documents", {})
    st.session_state.setdefault("uploaded_source_candidates", {})
    st.session_state.setdefault("uploaded_source_reports", {})
    st.session_state.setdefault("uploaded_issuer_data", {})
    st.session_state.setdefault("api_source_candidates", {})
    st.session_state.setdefault("api_source_reports", {})
    st.session_state.setdefault("manual_source_values", {})
    st.session_state.setdefault("approved_source_candidates", pd.DataFrame())

    required_names = list(_cached_required_names(methodology_id))
    required_df = _complete_required_field_frame(_required_field_frame(methodology_id), required_names)

    notice = st.session_state.pop("source_reset_notice", None)
    if notice:
        st.success(notice)

    top_cols = st.columns([1, 1, 2])
    with top_cols[0]:
        if st.button("Reset source session"):
            _reset_source_session(methodology_id)
            st.rerun()
    with top_cols[1]:
        guided_mode = st.checkbox(
            "Guided mode",
            value=bool(st.session_state.get("guided_source_mode", True)),
            key="guided_source_mode",
            help="Recommended while we pilot one issuer. Advanced API/manual tools stay available but collapsed.",
        )
    with top_cols[2]:
        st.caption("Reset only when changing issuer, methodology, or source workbook. Saved uploads stay in session until reset.")

    if guided_mode:
        st.info(
            "For this pilot: upload the CreditScope workbook, upload ACFR/debt PDFs, save issuer_data, then use the confirmation queue. "
            "Skip Census/BEA for now unless you specifically want to test API candidates."
        )
        st.dataframe(clean_for_display(_source_guided_progress()), width="stretch", hide_index=True)

    with st.expander("What each Source Data section does", expanded=False):
        st.dataframe(clean_for_display(pd.DataFrame(SOURCE_WORKFLOW_GUIDE)), width="stretch", hide_index=True)

    with st.container(border=True):
        st.markdown("**Step 1. Upload the files we have**" if guided_mode else "**Source uploads**")
        st.caption(
            "Start here. CreditScope can feed issuer_data; ACFR and debt PDFs are evidence until you approve specific values."
            if guided_mode
            else (
                "Upload source files for the Rating Path and Evidence Path. CreditScope workbooks can feed issuer_data; "
                "ACFR/OS PDFs are registered as evidence and will not change formula inputs until approved and applied."
            )
        )
        saved_uploads = _uploaded_sources_summary()
        if not saved_uploads.empty:
            st.success("Uploaded files are saved in this Streamlit session. Upload widgets may rerun the page, but these files remain available until Reset source session.")
            st.dataframe(clean_for_display(saved_uploads), width="stretch", hide_index=True)
        source_options = [
            {
                "key": "creditscope",
                "source_name": "CreditScope",
                "label": "CreditScope financial workbook",
                "types": ["csv", "xlsx", "xls"],
                "caption": "Required for this case. Upload Financials_ West Sacramento.xlsx here.",
                "multiple": False,
            },
            {
                "key": "ipeds",
                "source_name": "IPEDS_Excel",
                "label": "IPEDS Excel",
                "types": ["csv", "xlsx", "xls"],
                "caption": "CSV/XLSX/XLS up to 200MB per file.",
                "multiple": False,
            },
            {
                "key": "os",
                "source_name": "OS",
                "label": "Official Statement / debt support",
                "types": ["pdf", "csv", "xlsx", "xls"],
                "caption": "Upload one or more PDF/CSV/XLSX/XLS files. PDFs are evidence support and do not auto-replace scoring inputs.",
                "multiple": True,
            },
            {
                "key": "debt_report",
                "source_name": "DebtReport",
                "label": "Debt service / bonded indebtedness report",
                "types": ["pdf", "csv", "xlsx", "xls"],
                "caption": "Use this when no OS is available. Put CombinedDebtService / RemainingDebtService / BondedIndebtedness here.",
                "multiple": True,
            },
            {
                "key": "acfr",
                "source_name": "ACFR",
                "label": "ACFR / audited financial statements",
                "types": ["pdf", "csv", "xlsx", "xls"],
                "caption": "Upload ACFR PDFs here. They become evidence candidates, not automatic formula inputs.",
                "multiple": True,
            },
        ]
        if guided_mode:
            source_options = [
                option
                for option in source_options
                if option["key"] in {"creditscope", "acfr", "debt_report"}
            ]
        for row_start in range(0, len(source_options), 2):
            cols = st.columns(2)
            for col, option in zip(cols, source_options[row_start : row_start + 2]):
                key = option["key"]
                source_name = option["source_name"]
                label = option["label"]
                with col:
                    uploaded = st.file_uploader(
                        label,
                        type=option["types"],
                        key=f"upload_{key}",
                        help=option["caption"],
                        accept_multiple_files=bool(option.get("multiple", False)),
                    )
                    st.caption(option["caption"])
                    include_support_tabs = False
                    if source_name == "CreditScope" and not guided_mode:
                        include_support_tabs = st.checkbox(
                            "Include workbook support tabs as advanced supplemental source",
                            value=False,
                            key=f"include_support_tabs_{key}",
                            help=(
                                "Default off: only the auto-selected CreditScope/raw worksheet feeds issuer_data. "
                                "Turn on only when deliberately using scorecard/support tabs as a supplemental source."
                            ),
                        )
                    uploaded_files = uploaded if isinstance(uploaded, list) else ([uploaded] if uploaded is not None else [])
                    if not uploaded_files:
                        continue
                    try:
                        file_names: list[str] = []
                        source_data: dict[str, Any] = {}
                        report_frames: list[pd.DataFrame] = []
                        candidate_frames: list[pd.DataFrame] = []
                        pdf_count = 0

                        for uploaded_file in uploaded_files:
                            file_name, payload = _uploaded_file_payload(uploaded_file)
                            file_names.append(file_name)
                            if Path(file_name).suffix.lower() == ".pdf":
                                pdf_count += 1
                                _save_uploaded_pdf_document(key, source_name, file_name, payload)
                                continue

                            if source_name == "CreditScope":
                                sheet_names = _cached_excel_sheet_names(file_name, payload)
                                selected_sheet = _auto_sheet(sheet_names, file_name)
                                if sheet_names and selected_sheet is None:
                                    st.warning(
                                        "This does not look like a CreditScope raw workbook. If it is a debt/support workbook, "
                                        "upload it in Debt service / bonded indebtedness report instead."
                                    )
                                    st.caption(f"Detected worksheets: {', '.join(sheet_names)}")
                                    continue
                                if selected_sheet:
                                    st.success(f"Auto-selected: {selected_sheet}")
                                loader_output = _cached_creditscope_mapping(
                                    file_name,
                                    payload,
                                    selected_sheet,
                                    tuple(required_names),
                                    include_support_tabs,
                                    SOURCE_WORKFLOW_CACHE_VERSION,
                                )
                                report = loader_output["match_report"]
                                candidates = loader_output["source_candidates"]
                                mapped_data = dict(loader_output.get("issuer_data", {}) or {})
                            else:
                                mapped_data, report, candidates = _cached_mapped_upload(file_name, payload, source_name)

                            source_data.update(mapped_data)
                            if not report.empty and "uploaded_file" not in report.columns:
                                report.insert(0, "uploaded_file", file_name)
                            if isinstance(report, pd.DataFrame) and not report.empty:
                                report_frames.append(report)
                            if isinstance(candidates, pd.DataFrame) and not candidates.empty:
                                candidate_frames.append(candidates)

                        report = pd.concat(report_frames, ignore_index=True) if report_frames else pd.DataFrame()
                        candidates = (
                            pd.concat(candidate_frames, ignore_index=True)
                            if candidate_frames
                            else pd.DataFrame()
                        )
                        file_label = "; ".join(file_names)
                        st.session_state["uploaded_sources"][key] = file_label
                        st.session_state["uploaded_source_candidates"][key] = candidates
                        st.session_state["uploaded_source_reports"][key] = report
                        st.session_state["uploaded_issuer_data"][key] = {
                            "source_name": source_name,
                            "file_name": file_label,
                            "issuer_data": source_data,
                        }
                        if source_data and pdf_count:
                            st.success(f"{len(source_data)} fields mapped; {pdf_count} PDF file(s) registered for Source QA.")
                        elif source_data:
                            st.success(f"{len(source_data)} fields mapped across {len(file_names)} file(s).")
                        elif pdf_count:
                            st.success(f"{pdf_count} PDF file(s) registered for Source QA.")
                        else:
                            st.info("Files uploaded, but no mappable fields were found.")
                    except Exception as exc:
                        st.error("Could not map uploaded file(s).")
                        st.exception(exc)

        upload_reports = [
            frame for frame in st.session_state.get("uploaded_source_reports", {}).values() if isinstance(frame, pd.DataFrame) and not frame.empty
        ]
        if upload_reports:
            with st.expander("Upload diagnostics", expanded=False):
                st.caption("This shows what each upload could and could not map. It is not the final source readiness list.")
                st.dataframe(clean_for_display(pd.concat(upload_reports, ignore_index=True)), width="stretch", hide_index=True)

    api_container = st.expander("Step 2. API source candidates (optional)", expanded=True)
    with api_container:
        defaults = _recommended_api_defaults()
        _set_recommended_api_widget_defaults(defaults)
        st.markdown("**API source candidates**")
        st.caption("These values become source candidates only. They do not enter formulas until they appear in and are saved from the issuer_data table.")
        status_cols = st.columns(4)
        status_cols[0].metric("Scorecard Year", defaults["analysis_year_label"])
        status_cols[1].metric("ACS Source Year", defaults["census_year"])
        status_cols[2].metric("BEA Current", defaults["bea_year"])
        status_cols[3].metric("BEA Prior", defaults["bea_prior"])
        st.info(
            "Why 2024? The rating/scorecard year can be 2026, but public Census ACS and BEA county-level API data "
            "lag the analysis year. This panel uses the latest supported public source vintage as API evidence; "
            "newer issuer-specific values can still be entered in the issuer_data table from CreditScope, ACFR, OS, "
            "DebtReport, or manual confirmation."
        )
        if st.button("Use recommended API defaults", key="api_use_recommended_defaults"):
            _set_recommended_api_widget_defaults(defaults, force=True)
            st.rerun()

        geo_cols = st.columns([1, 1, 2])
        state_fips = geo_cols[0].text_input(
            "State FIPS",
            max_chars=2,
            key="api_state_fips",
            help="Used by both Census ACS and BEA.",
        )
        county_fips = geo_cols[1].text_input(
            "County FIPS",
            max_chars=3,
            key="api_county_fips",
            help="Used by both Census ACS and BEA.",
        )
        if defaults.get("is_west_sacramento"):
            geo_cols[2].info("West Sacramento should use California 06 and Yolo County 113.")
        else:
            geo_cols[2].caption("Confirm county FIPS before fetching API candidates.")
        st.session_state["state_fips"] = str(state_fips).zfill(2)
        st.session_state["county_fips"] = str(county_fips).zfill(3)

        c1, c2 = st.columns(2)
        with c1:
            st.write("Census ACS")
            census_key_available = bool(get_census_api_key())
            if not census_key_available:
                _show_missing_api_key("Census ACS", "CENSUS_API_KEY")
            census_year = st.number_input(
                "ACS source year",
                min_value=2009,
                max_value=LATEST_CENSUS_SOURCE_YEAR,
                step=1,
                key="api_census_source_year",
                help="This is the Census data vintage, not the scorecard year.",
            )
            include_proxy = st.checkbox("Include proxy fields", value=False)
            census_fields = supported_census_candidate_fields(include_proxy_fields=include_proxy)
            st.caption(f"{len(census_fields)} Census fields selected automatically.")
            if st.button("Fetch Census", key="api_fetch_census", disabled=not census_key_available):
                try:
                    census = _cached_census_source_candidates(
                        str(state_fips),
                        str(county_fips),
                        int(census_year),
                        tuple(census_fields),
                        bool(include_proxy),
                    )
                    st.session_state["api_source_candidates"]["census"] = census
                    st.session_state["api_source_reports"]["census"] = census
                    st.success(f"Fetched {len(census)} Census candidate fields.")
                except CensusApiError as exc:
                    _show_api_fetch_error("Census ACS", exc)
        with c2:
            st.write("BEA Regional")
            bea_key_available = bool(get_bea_api_key())
            if not bea_key_available:
                _show_missing_api_key("BEA Regional", "BEA_API_KEY")
            bea_cols = st.columns(2)
            bea_year = bea_cols[0].number_input(
                "BEA source year",
                min_value=2001,
                max_value=LATEST_BEA_SOURCE_YEAR,
                step=1,
                key="api_bea_source_year",
                help="This is the BEA data vintage, not the scorecard year.",
            )
            bea_prior = max(2000, int(bea_year) - 1)
            bea_cols[1].text_input(
                "BEA prior year",
                value=str(bea_prior),
                disabled=True,
                help="Automatically set to the year before the BEA source year.",
            )
            bea_fields = supported_bea_candidate_fields()
            st.caption(f"{len(bea_fields)} BEA fields selected automatically.")
            if st.button("Fetch BEA", key="api_fetch_bea", disabled=not bea_key_available):
                try:
                    bea = _cached_bea_source_candidates(
                        str(state_fips),
                        str(county_fips),
                        int(bea_year),
                        int(bea_prior),
                        tuple(bea_fields),
                    )
                    st.session_state["api_source_candidates"]["bea"] = bea
                    st.session_state["api_source_reports"]["bea"] = bea
                    st.success(f"Fetched {len(bea)} BEA candidate fields.")
                except BeaApiError as exc:
                    _show_api_fetch_error("BEA", exc)
        can_fetch_any_api = census_key_available or bea_key_available
        if st.button("Fetch Census + BEA", key="api_fetch_all", type="primary", disabled=not can_fetch_any_api):
            fetched: list[str] = []
            if census_key_available:
                try:
                    census = _cached_census_source_candidates(
                        str(state_fips),
                        str(county_fips),
                        int(census_year),
                        tuple(census_fields),
                        bool(include_proxy),
                    )
                    st.session_state["api_source_candidates"]["census"] = census
                    st.session_state["api_source_reports"]["census"] = census
                    fetched.append(f"{len(census)} Census")
                except CensusApiError as exc:
                    _show_api_fetch_error("Census ACS", exc)
            if bea_key_available:
                try:
                    bea = _cached_bea_source_candidates(
                        str(state_fips),
                        str(county_fips),
                        int(bea_year),
                        int(bea_prior),
                        tuple(bea_fields),
                    )
                    st.session_state["api_source_candidates"]["bea"] = bea
                    st.session_state["api_source_reports"]["bea"] = bea
                    fetched.append(f"{len(bea)} BEA")
                except BeaApiError as exc:
                    _show_api_fetch_error("BEA", exc)
            if fetched:
                st.success(f"Fetched {' and '.join(fetched)} candidate fields.")

    with st.container(border=True):
        st.markdown("**Step 3. Save issuer_data**" if guided_mode else "**Save issuer_data**")
        st.caption(
            "This is the official calculation input table. Values found from uploads/API/approved sources are prefilled; "
            "required raw fields that are not found stay blank. Derived ratios are calculated later in Formula Results, "
            "not typed here. Type manual replacements directly in the value column."
        )
        frames = _source_candidate_frames(include_manual_values=False)
        issuer_data_editor, source_context = _build_issuer_data_editor(
            methodology_id=methodology_id,
            required_fields=required_df,
            required_names=required_names,
            frames=frames,
        )
        if frames:
            st.caption(f"{len(frames)} source candidate group(s) are feeding the suggested values in this table.")
        else:
            st.info("No source candidates yet. The table is still available so required calculation inputs are visible.")
        with st.form(f"issuer_data_input_form_{methodology_id}"):
            edited_inputs = st.data_editor(
                issuer_data_editor,
                width="stretch",
                hide_index=True,
                num_rows="fixed",
                key=f"issuer_data_input_editor_{methodology_id}",
                column_config={
                    "field_name": st.column_config.TextColumn("field_name", disabled=True),
                    "value": st.column_config.TextColumn("value"),
                    "source_status": st.column_config.TextColumn("source_status", disabled=True),
                    "source_used": st.column_config.TextColumn("source_used", disabled=True),
                    "used_by": st.column_config.TextColumn("used_by", disabled=True),
                    "category": st.column_config.TextColumn("category", disabled=True),
                },
            )
            button_cols = st.columns(2)
            save_inputs = button_cols[0].form_submit_button(
                "Save issuer_data",
                type="primary",
            )
            save_inputs_and_run = button_cols[1].form_submit_button("Save issuer_data and run formulas")

        if save_inputs or save_inputs_and_run:
            _save_issuer_data_from_editor(
                edited_inputs,
                methodology_id=methodology_id,
                source_context=source_context,
                run_formulas=save_inputs_and_run,
            )

    with st.expander("Support inventory readiness (advanced, not rating blockers)", expanded=False):
        _readiness_tabs(st.session_state.get("source_report", pd.DataFrame()), methodology_id)
