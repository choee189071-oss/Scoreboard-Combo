from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_OUTPUT = Path(
    "/Users/zhouyiyi/Documents/Codex/2026-06-17/"
    "https-combinationofscoreboard-streamlit-app-streamlit-app/outputs/"
    "methodology_accuracy_matrix.xlsx"
)

TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4D6")
GRAY_FILL = PatternFill("solid", fgColor="EDEDED")
BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
TABLE_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def _clean_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    return json.dumps(value, ensure_ascii=False)


def _as_rows(rows: list[dict[str, Any]], columns: list[str]) -> list[list[Any]]:
    return [[_clean_cell_value(row.get(column)) for column in columns] for row in rows]


def _write_title(ws, title: str, notes: Iterable[str] = ()) -> int:
    ws["A1"] = title
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    row = 3
    for note in notes:
        ws.cell(row=row, column=1, value=note)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row + 1 if row > 3 else 3


def _autosize(ws, columns: list[str], start_row: int, max_width: int = 48) -> None:
    for idx, column in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        values = [str(column)]
        for row in range(start_row + 1, min(ws.max_row, start_row + 80) + 1):
            values.append(str(ws.cell(row=row, column=idx).value or ""))
        width = max(len(value) for value in values)
        ws.column_dimensions[letter].width = min(max(width + 2, 11), max_width)


def _style_table(ws, start_row: int, columns: list[str], table_name: str) -> None:
    if not columns:
        return
    end_col = get_column_letter(len(columns))
    end_row = max(ws.max_row, start_row)
    header_range = f"A{start_row}:{end_col}{start_row}"
    for row in ws[header_range]:
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = TABLE_BORDER

    for row in ws.iter_rows(min_row=start_row + 1, max_row=end_row, max_col=len(columns)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = TABLE_BORDER

    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{end_col}{end_row}"

    if end_row > start_row:
        table = Table(displayName=table_name, ref=f"A{start_row}:{end_col}{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)


def _fill_row(ws, row_number: int, fill: PatternFill, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row=row_number, column=col).fill = fill


def _status_fill(row: dict[str, Any]) -> PatternFill | None:
    value_status = str(row.get("value_status", "") or "").lower()
    candidate_status = str(row.get("candidate_status", "") or "").lower()
    raw_status = str(row.get("raw_status", "") or "").lower()
    official_match = str(row.get("official_match", "") or "").lower()
    blocking = str(row.get("blocking", "") or "").lower()

    if blocking == "true":
        return RED_FILL
    if value_status == "match" or candidate_status in {"match", "support_primary_match"} or official_match == "true":
        return GREEN_FILL
    if value_status in {"mismatch", "model_missing"} or "mismatch" in candidate_status:
        return YELLOW_FILL
    if raw_status and raw_status != "ok":
        return GRAY_FILL
    if value_status == "manual_skip":
        return BLUE_FILL
    return None


def _add_table_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    columns: list[str],
    table_name: str,
    notes: Iterable[str] = (),
) -> None:
    ws = wb.create_sheet(title[:31])
    start_row = _write_title(ws, title, notes)
    for col_idx, column in enumerate(columns, start=1):
        ws.cell(row=start_row, column=col_idx, value=column)
    for row_idx, values in enumerate(_as_rows(rows, columns), start=start_row + 1):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for offset, row in enumerate(rows, start=start_row + 1):
        fill = _status_fill(row)
        if fill is not None:
            _fill_row(ws, offset, fill, len(columns))

    _style_table(ws, start_row, columns, table_name)
    _autosize(ws, columns, start_row)


def _status_counts(rows: list[dict[str, Any]], column: str) -> list[dict[str, Any]]:
    return [{"status": status, "count": count} for status, count in Counter(str(row.get(column) or "blank") for row in rows).most_common()]


def _add_cover(wb: Workbook, tables: dict[str, list[dict[str, Any]]], manifest: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = "Methodology Accuracy Matrix"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].fill = TITLE_FILL
    ws.row_dimensions[1].height = 28

    summary = tables.get("summary", [])
    accuracy = tables.get("accuracy_matrix", [])
    coverage = tables.get("field_coverage", [])
    page = tables.get("page_consistency", [])
    blocking = sum(1 for row in accuracy if str(row.get("blocking")).lower() == "true")
    raw_ok = sum(1 for row in summary if row.get("raw_status") == "ok")
    not_run = len(summary) - raw_ok
    value_counts = Counter(str(row.get("value_status") or "blank") for row in accuracy)

    rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("No-cheat rule", manifest.get("rule", "")),
        ("Cases", len(summary)),
        ("Cases with primary raw/CreditScope sheet", raw_ok),
        ("Cases without primary raw/CreditScope sheet", not_run),
        ("Accuracy matrix rows", len(accuracy)),
        ("Field coverage rows", len(coverage)),
        ("Blocking metric rows", blocking),
        ("Page consistency rows", len(page)),
        ("Value status counts", ", ".join(f"{k}: {v}" for k, v in value_counts.most_common())),
        ("", ""),
        ("Key takeaways", ""),
        ("1", "S&P Local Gov support tabs contain direct metric values that match official scorecard metrics, but those values are evidence checks only."),
        ("2", "Contra Costa Moody's workbook contains a non-matching Cerritos CCD FIN tab; the run uses the Contra Costa raw tab."),
        ("3", "The largest current gap is source coverage and normalization from CreditScope/raw sheets, not only score-threshold wiring."),
    ]

    for row_idx, (label, value) in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=1).font = BOLD_FONT
        ws.cell(row=row_idx, column=1).border = TABLE_BORDER
        ws.cell(row=row_idx, column=2).border = TABLE_BORDER
        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        if label == "Key takeaways":
            ws.cell(row=row_idx, column=1).fill = SUBHEADER_FILL
            ws.cell(row=row_idx, column=2).fill = SUBHEADER_FILL

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A3"


def build_workbook(data_dir: Path, output_path: Path) -> Path:
    tables_path = data_dir / "tables.json"
    manifest_path = data_dir / "manifest.json"
    tables: dict[str, list[dict[str, Any]]] = json.loads(tables_path.read_text(encoding="utf-8"))
    manifest: dict[str, Any] = json.loads(manifest_path.read_text(encoding="utf-8"))

    wb = Workbook()
    _add_cover(wb, tables, manifest)
    _add_table_sheet(
        wb,
        "Summary",
        tables.get("summary", []),
        [
            "fixture_key",
            "methodology_id",
            "issuer_name",
            "workbook",
            "primary_raw_sheet",
            "raw_status",
            "primary_input_fields",
            "formula_rows",
            "official_metric_rows",
            "value_matches",
            "value_mismatches",
            "model_missing",
            "manual_skips",
            "model_rating",
            "official_rating",
            "model_weighted_score",
            "official_weighted_score",
            "coverage_status",
            "notes",
        ],
        "SummaryTable",
        [
            "Primary model values are generated from the configured raw/CreditScope sheet only.",
            "Cases marked no_primary_raw_sheet are audited for workbook consistency but not run through the primary model path.",
        ],
    )
    _add_table_sheet(
        wb,
        "Accuracy Matrix",
        tables.get("accuracy_matrix", []),
        [
            "fixture_key",
            "methodology_id",
            "issuer_name",
            "factor",
            "metric",
            "formula_id",
            "official_weight",
            "official_value",
            "model_value",
            "model_compare_value",
            "value_delta",
            "value_status",
            "official_score",
            "model_score",
            "score_delta",
            "score_match",
            "formula_status",
            "required_fields",
            "raw_source_cells",
            "missing_fields",
            "warning",
            "blocking",
            "suspected_cause",
        ],
        "AccuracyMatrixTable",
        ["Each row compares one official scorecard metric with the no-cheat model output."],
    )
    _add_table_sheet(
        wb,
        "Primary Raw Inputs",
        tables.get("primary_raw_inputs", []),
        [
            "test_case",
            "methodology_id",
            "issuer_name",
            "field_name",
            "value",
            "source_workbook",
            "source_sheet",
            "source_cell",
            "source_label",
            "source_type",
            "notes",
        ],
        "PrimaryRawInputsTable",
    )
    _add_table_sheet(
        wb,
        "Field Coverage",
        tables.get("field_coverage", []),
        [
            "fixture_key",
            "methodology_id",
            "issuer_name",
            "formula_id",
            "metric",
            "official_weight",
            "formula_value_status",
            "metric_blocking",
            "field_name",
            "field_category",
            "coverage_status",
            "primary_value",
            "source_sheet",
            "source_cell",
            "source_label",
            "source_type",
            "no_cheat_allowed",
            "preferred_source",
            "fallback_source",
            "likely_sources",
            "other_page_candidate_statuses",
            "other_page_evidence_sheets",
            "other_page_official_match",
            "field_blocking",
            "suspected_cause",
        ],
        "FieldCoverageTable",
        ["Primary coverage only counts configured CreditScope/raw tabs; other page evidence is shown separately."],
    )
    _add_table_sheet(
        wb,
        "Page Consistency",
        tables.get("page_consistency", []),
        [
            "fixture_key",
            "methodology_id",
            "issuer_name",
            "workbook",
            "primary_raw_sheet",
            "sheet_name",
            "field_name",
            "matched_label",
            "candidate_cell",
            "candidate_value",
            "primary_value",
            "delta_to_primary",
            "official_value",
            "delta_to_official",
            "official_match",
            "candidate_status",
            "match_method",
            "notes",
        ],
        "PageConsistencyTable",
        ["Other workbook sheets are evidence checks only; they are not fed into the no-cheat primary model."],
    )
    _add_table_sheet(
        wb,
        "Source Quality",
        tables.get("source_quality", []),
        ["fixture_key", "source_category", "source_type", "raw_field_count"],
        "SourceQualityTable",
    )
    _add_table_sheet(
        wb,
        "Value Status Counts",
        _status_counts(tables.get("accuracy_matrix", []), "value_status"),
        ["status", "count"],
        "ValueStatusCountsTable",
    )
    _add_table_sheet(
        wb,
        "Manifest",
        [
            {
                "created_by": manifest.get("created_by"),
                "rule": manifest.get("rule"),
                "row_counts": json.dumps(manifest.get("row_counts", {}), ensure_ascii=False),
            }
        ],
        ["created_by", "rule", "row_counts"],
        "ManifestTable",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    check = load_workbook(output_path, read_only=True, data_only=True)
    try:
        if "Accuracy Matrix" not in check.sheetnames:
            raise RuntimeError("Workbook verification failed: Accuracy Matrix sheet missing.")
    finally:
        check.close()
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the Methodology Accuracy Matrix Excel workbook.")
    parser.add_argument("--data-dir", default="work/methodology_accuracy_matrix")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    print(build_workbook(Path(args.data_dir), Path(args.output)))


if __name__ == "__main__":
    main()
