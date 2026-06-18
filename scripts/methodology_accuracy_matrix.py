from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from connectors.creditscope_loader import _load_sp_local_gov_supplemental_report, load_creditscope_source_candidates
from engine.calculator_engine import load_formula_library, parse_required_fields
from engine.mapping_engine import map_creditscope_workbook
from engine.raw_value_validation import (
    calculate_template_formulas_from_raw,
    clean_optional_float,
    compare_auto_scores_to_official,
    compare_raw_formula_values_to_official,
    raw_fixture_to_issuer_data,
    raw_source_quality_summary,
)
from engine.official_fixture_engine import load_official_fixture
from engine.rating_engine import run_rating_engine


DOWNLOAD_DIR = Path("/Users/zhouyiyi/Downloads/Ramirez 2")


@dataclass(frozen=True)
class ScorecardCase:
    fixture_key: str
    methodology_id: str
    issuer_name: str
    workbook_path: Path
    primary_raw_sheet: str | None
    notes: str = ""


CASES: tuple[ScorecardCase, ...] = (
    ScorecardCase(
        fixture_key="west_sacramento_sp_local_gov_k12",
        methodology_id="sp_local_gov_k12",
        issuer_name="City of West Sacramento",
        workbook_path=DOWNLOAD_DIR / "City of West Sacramento - Local Govt Scorecard 2024 (5-22-26).xlsx",
        primary_raw_sheet="CreditScope",
    ),
    ScorecardCase(
        fixture_key="contra_costa_sp_community_college_go",
        methodology_id="sp_community_college_go",
        issuer_name="Contra Costa CCD",
        workbook_path=DOWNLOAD_DIR / "Contra Costa CCD GO Scorecard 2023.xlsx",
        primary_raw_sheet="Contra Costa CCD",
    ),
    ScorecardCase(
        fixture_key="contra_costa_moodys_ccd_go",
        methodology_id="moodys_ccd_go",
        issuer_name="Contra Costa Community College District",
        workbook_path=DOWNLOAD_DIR / "Contra Costa CCD Moodys Higher Ed Scorecard.xlsx",
        primary_raw_sheet="Contra Costa",
        notes="Issuer-aware override: the generic picker prefers Cerritos CCD FIN, which is a non-matching raw tab.",
    ),
    ScorecardCase(
        fixture_key="alum_rock_moodys_k12",
        methodology_id="moodys_k12",
        issuer_name="Alum Rock Union ESD",
        workbook_path=DOWNLOAD_DIR / "Moody's K-12 Scorecard - Alum Rock Union ESD (3-24-26).xlsx",
        primary_raw_sheet="Credit Scope",
    ),
    ScorecardCase(
        fixture_key="jefferson_sp_local_gov_k12",
        methodology_id="sp_local_gov_k12",
        issuer_name="Jefferson Union High School District",
        workbook_path=DOWNLOAD_DIR / "SP Local Govt Scorecard 2024 - Jefferson UHSD USD (3-11-26).xlsx",
        primary_raw_sheet=None,
        notes="No distinct CreditScope/raw page is present in the workbook; support tabs are audited separately.",
    ),
    ScorecardCase(
        fixture_key="ontario_sp_water_sewer",
        methodology_id="sp_water_sewer",
        issuer_name="City of Ontario Water/Sewer",
        workbook_path=DOWNLOAD_DIR / "SP Water & Sewer Utility Scorecards - City of Ontario - 7-29-25.xlsx",
        primary_raw_sheet=None,
        notes="No distinct CreditScope/raw page is present in the workbook; scorecard pages are audited separately.",
    ),
)


def _safe_float(value: Any) -> float | None:
    number = clean_optional_float(value)
    return None if number is None or pd.isna(number) else float(number)


def _json_default(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    return str(value)


def _clean_json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(k): _clean_json_value(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_clean_json_value(item) for item in value]
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def _records_for_json(frame: pd.DataFrame) -> list[dict[str, Any]]:
    return [
        {str(key): _clean_json_value(value) for key, value in record.items()}
        for record in frame.to_dict(orient="records")
    ]


def _load_sheet_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _raw_fixture_from_primary(case: ScorecardCase) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    if case.primary_raw_sheet is None:
        return pd.DataFrame(), pd.DataFrame(), "no_primary_raw_sheet"
    if not case.workbook_path.exists():
        return pd.DataFrame(), pd.DataFrame(), "workbook_missing"

    sheet_names = _load_sheet_names(case.workbook_path)
    if case.primary_raw_sheet not in sheet_names:
        return pd.DataFrame(), pd.DataFrame(), "primary_raw_sheet_not_found"

    extracted = load_creditscope_source_candidates(
        case.workbook_path,
        sheet_name=case.primary_raw_sheet,
        include_support_tabs=False,
    )
    issuer_data = extracted.get("issuer_data", {}) or {}
    report = extracted.get("match_report", pd.DataFrame())
    if not issuer_data:
        return pd.DataFrame(), report if isinstance(report, pd.DataFrame) else pd.DataFrame(), "no_mappable_raw_fields"

    report_lookup: dict[str, dict[str, Any]] = {}
    if isinstance(report, pd.DataFrame) and not report.empty and "field_name" in report.columns:
        for _, row in report.iterrows():
            field = str(row.get("field_name", "") or "").strip()
            if field and field not in report_lookup:
                report_lookup[field] = row.to_dict()

    rows = []
    for field_name, value in sorted(issuer_data.items()):
        source = report_lookup.get(field_name, {})
        rows.append(
            {
                "test_case": case.fixture_key,
                "methodology_id": case.methodology_id,
                "issuer_name": case.issuer_name,
                "field_name": field_name,
                "value": value,
                "unit": "",
                "source_workbook": case.workbook_path.name,
                "source_sheet": source.get("sheet_name", case.primary_raw_sheet),
                "source_cell": source.get("matched_column", ""),
                "source_label": source.get("matched_label", ""),
                "source_type": "creditscope_raw_primary",
                "notes": "Primary no-cheat input from configured raw/CreditScope sheet.",
            }
        )
    return pd.DataFrame(rows), report if isinstance(report, pd.DataFrame) else pd.DataFrame(), "ok"


def _source_summary_for_compare(compare: pd.DataFrame) -> dict[str, int]:
    if compare.empty or "value_status" not in compare.columns:
        return {}
    return {str(k): int(v) for k, v in compare["value_status"].fillna("unknown").value_counts().to_dict().items()}


def _status_counts(df: pd.DataFrame, column: str) -> dict[str, int]:
    if df.empty or column not in df.columns:
        return {}
    return {str(k): int(v) for k, v in df[column].fillna("unknown").value_counts().to_dict().items()}


def _reason_from_row(row: pd.Series, raw_status: str) -> str:
    value_status = str(row.get("value_status", "") or "")
    formula_status = str(row.get("status", "") or "")
    missing_fields = str(row.get("missing_fields", "") or "").strip()
    warning = str(row.get("warning", "") or "").strip()
    if raw_status != "ok":
        return raw_status
    if value_status == "match":
        return "matches official benchmark"
    if value_status == "manual_skip":
        return "manual or qualitative score; excluded from formula accuracy"
    if value_status == "model_missing":
        return f"missing raw field(s): {missing_fields or 'not reported'}"
    if value_status == "source_pending":
        return "source was not independent in the raw fixture"
    if value_status == "mismatch":
        return warning or "formula, unit, denominator, or period mismatch"
    if formula_status in {"missing", "error"}:
        return f"formula status {formula_status}"
    return value_status or "needs review"


def _accuracy_rows(
    case: ScorecardCase,
    official: pd.DataFrame,
    value_comparison: pd.DataFrame,
    score_comparison: pd.DataFrame,
    raw_status: str,
) -> pd.DataFrame:
    if official.empty:
        return pd.DataFrame()
    value = value_comparison.copy() if not value_comparison.empty else official.copy()
    if "formula_id" not in value.columns:
        return pd.DataFrame()

    score = score_comparison.copy()
    score_cols = [
        "formula_id",
        "raw_value",
        "model_score",
        "score_label",
        "score_delta",
        "score_match",
        "missing_reason",
    ]
    if not score.empty and "formula_id" in score.columns:
        score = score[[col for col in score_cols if col in score.columns]]
        merged = value.merge(score, on="formula_id", how="left", suffixes=("", "_score_model"))
    else:
        merged = value.copy()

    rows = []
    for _, row in merged.iterrows():
        official_weight = _safe_float(row.get("official_weight"))
        value_status = str(row.get("value_status", "") or "not_run")
        score_match = row.get("score_match")
        is_blocking = bool(
            (official_weight or 0) > 0
            and value_status not in {"match", "manual_skip", "official_missing"}
        )
        rows.append(
            {
                "fixture_key": case.fixture_key,
                "methodology_id": case.methodology_id,
                "issuer_name": case.issuer_name,
                "workbook": case.workbook_path.name,
                "primary_raw_sheet": case.primary_raw_sheet or "",
                "raw_status": raw_status,
                "section": row.get("section", ""),
                "factor": row.get("factor", ""),
                "metric": row.get("metric", ""),
                "formula_id": row.get("formula_id", ""),
                "official_weight": official_weight,
                "official_value": row.get("official_value", ""),
                "model_value": row.get("model_value", ""),
                "model_compare_value": row.get("model_compare_value", ""),
                "value_delta": row.get("value_delta", ""),
                "value_status": value_status,
                "official_score": row.get("official_score", ""),
                "model_score": row.get("model_score", ""),
                "score_delta": row.get("score_delta", ""),
                "score_match": score_match,
                "formula_status": row.get("status", ""),
                "required_fields": row.get("required_fields", ""),
                "raw_source_cells": row.get("raw_source_cells", ""),
                "missing_fields": row.get("missing_fields", ""),
                "warning": row.get("warning", ""),
                "blocking": is_blocking,
                "suspected_cause": _reason_from_row(row, raw_status),
            }
        )
    return pd.DataFrame(rows)


def _run_rating(case: ScorecardCase, formula_results: pd.DataFrame) -> tuple[dict[str, Any], pd.DataFrame]:
    if formula_results.empty:
        return {}, pd.DataFrame()
    output = run_rating_engine(
        methodology_id=case.methodology_id,
        formula_results=formula_results,
        manual_scores={},
        thresholds_path="config/scoring_thresholds.csv",
        templates_dir="templates",
    )
    metric_scores = output.get("factor_engine_output", {}).get("metric_scores", pd.DataFrame())
    return output, metric_scores if isinstance(metric_scores, pd.DataFrame) else pd.DataFrame()


def _rating_result(output: dict[str, Any]) -> dict[str, Any]:
    result = output.get("rating_result", {}) if isinstance(output, dict) else {}
    return result if isinstance(result, dict) else {}


def _official_value_lookup(official: pd.DataFrame) -> dict[str, float | None]:
    if official.empty or "formula_id" not in official.columns:
        return {}
    return {
        str(row.get("formula_id", "") or "").strip(): _safe_float(row.get("official_value"))
        for _, row in official.iterrows()
    }


def _field_metadata_lookup(
    data_dictionary_path: str | Path = "config/data_dictionary.csv",
    field_mapping_path: str | Path = "config/field_mapping.csv",
) -> dict[str, dict[str, str]]:
    metadata: dict[str, dict[str, str]] = {}
    dictionary_path = Path(data_dictionary_path)
    if dictionary_path.exists():
        dictionary = pd.read_csv(dictionary_path)
        for _, row in dictionary.iterrows():
            field = str(row.get("field_name", "") or "").strip()
            if not field:
                continue
            preferred = str(row.get("preferred_source", "") or "").strip()
            fallback = str(row.get("fallback_source", "") or "").strip()
            sources = [
                part.strip()
                for value in [preferred, fallback]
                for part in value.replace(";", "|").split("|")
                if part.strip() and part.strip().lower() != "nan"
            ]
            metadata[field] = {
                "field_category": str(row.get("field_category", "") or "").strip(),
                "preferred_source": preferred,
                "fallback_source": fallback,
                "likely_sources": "|".join(dict.fromkeys(sources)),
            }

    mapping_path = Path(field_mapping_path)
    if mapping_path.exists():
        mapping = pd.read_csv(mapping_path)
        if {"field_name", "source_name"} <= set(mapping.columns):
            grouped = (
                mapping.assign(
                    field_name=mapping["field_name"].astype(str).str.strip(),
                    source_name=mapping["source_name"].astype(str).str.strip(),
                )
                .groupby("field_name")["source_name"]
                .apply(lambda s: "|".join(sorted(set(x for x in s if x and x.lower() != "nan"))))
                .to_dict()
            )
            for field, sources in grouped.items():
                item = metadata.setdefault(
                    field,
                    {
                        "field_category": "",
                        "preferred_source": "",
                        "fallback_source": "",
                        "likely_sources": "",
                    },
                )
                if sources:
                    item["likely_sources"] = "|".join(
                        dict.fromkeys(
                            [
                                part
                                for value in [item.get("likely_sources", ""), sources]
                                for part in str(value).split("|")
                                if part
                            ]
                        )
                    )
    return metadata


def _required_fields_lookup(formula_library_path: str | Path = "config/formula_library.csv") -> dict[str, list[str]]:
    library = load_formula_library(formula_library_path)
    return {
        str(row.get("formula_id", "") or "").strip(): parse_required_fields(row.get("required_data", ""))
        for _, row in library.iterrows()
        if str(row.get("formula_id", "") or "").strip()
    }


def _raw_lookup(primary_raw: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if primary_raw.empty or "field_name" not in primary_raw.columns:
        return {}
    out: dict[str, dict[str, Any]] = {}
    for _, row in primary_raw.iterrows():
        field = str(row.get("field_name", "") or "").strip()
        if field and field not in out:
            out[field] = row.to_dict()
    return out


def _accuracy_lookup(accuracy: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if accuracy.empty or "formula_id" not in accuracy.columns:
        return {}
    out: dict[str, dict[str, Any]] = {}
    for _, row in accuracy.iterrows():
        formula_id = str(row.get("formula_id", "") or "").strip()
        if formula_id and formula_id not in out:
            out[formula_id] = row.to_dict()
    return out


def _other_page_evidence(pages: pd.DataFrame, field_name: str) -> dict[str, Any]:
    if pages.empty or "field_name" not in pages.columns:
        return {
            "other_page_candidate_statuses": "",
            "other_page_evidence_sheets": "",
            "other_page_official_match": None,
        }
    matches = pages[pages["field_name"].astype(str).str.strip().eq(field_name)].copy()
    if matches.empty:
        return {
            "other_page_candidate_statuses": "",
            "other_page_evidence_sheets": "",
            "other_page_official_match": None,
        }
    statuses = "|".join(sorted(set(matches.get("candidate_status", pd.Series(dtype=str)).dropna().astype(str))))
    sheets = "|".join(sorted(set(matches.get("sheet_name", pd.Series(dtype=str)).dropna().astype(str))))
    official_match_values = matches.get("official_match", pd.Series(dtype=object)).dropna().astype(str).str.lower()
    official_match = None if official_match_values.empty else bool(official_match_values.eq("true").any())
    return {
        "other_page_candidate_statuses": statuses,
        "other_page_evidence_sheets": sheets,
        "other_page_official_match": official_match,
    }


def _field_coverage_rows(
    case: ScorecardCase,
    official: pd.DataFrame,
    primary_raw: pd.DataFrame,
    raw_status: str,
    accuracy: pd.DataFrame,
    pages: pd.DataFrame,
) -> pd.DataFrame:
    if official.empty or "formula_id" not in official.columns:
        return pd.DataFrame()

    required_lookup = _required_fields_lookup()
    metadata_lookup = _field_metadata_lookup()
    raw_by_field = _raw_lookup(primary_raw)
    accuracy_by_formula = _accuracy_lookup(accuracy)
    rows: list[dict[str, Any]] = []

    for _, metric_row in official.iterrows():
        formula_id = str(metric_row.get("formula_id", "") or "").strip()
        metric_accuracy = accuracy_by_formula.get(formula_id, {})
        official_weight = _safe_float(metric_row.get("official_weight")) or 0.0
        metric_blocking = str(metric_accuracy.get("blocking", "")).lower() == "true"
        formula_known = formula_id in required_lookup
        fields = required_lookup.get(formula_id, [])

        if not formula_known:
            field_names = ["formula_definition"]
        elif not fields:
            field_names = ["no_raw_field_required"]
        elif fields == ["manual"]:
            field_names = ["manual_score"]
        else:
            field_names = fields

        for field_name in field_names:
            raw_row = raw_by_field.get(field_name, {})
            raw_value = raw_row.get("value", "")
            has_primary_value = bool(raw_row) and str(raw_value).strip() != "" and str(raw_value).lower() != "nan"

            if field_name == "formula_definition":
                coverage_status = "formula_missing"
            elif field_name == "manual_score":
                coverage_status = "manual_score"
            elif field_name == "no_raw_field_required":
                coverage_status = "not_applicable"
            elif raw_status != "ok":
                coverage_status = raw_status
            elif has_primary_value:
                coverage_status = "available_primary_raw"
            else:
                coverage_status = "missing_primary_raw"

            source_type = str(raw_row.get("source_type", "") or "")
            no_cheat_allowed = coverage_status == "available_primary_raw" and "creditscope_raw_primary" in source_type
            metadata = metadata_lookup.get(field_name, {})
            other_evidence = _other_page_evidence(pages, field_name)
            field_blocking = bool(
                metric_blocking
                and official_weight > 0
                and coverage_status
                not in {
                    "manual_score",
                    "not_applicable",
                }
            )

            rows.append(
                {
                    "fixture_key": case.fixture_key,
                    "methodology_id": case.methodology_id,
                    "issuer_name": case.issuer_name,
                    "workbook": case.workbook_path.name,
                    "primary_raw_sheet": case.primary_raw_sheet or "",
                    "raw_status": raw_status,
                    "section": metric_row.get("section", ""),
                    "factor": metric_row.get("factor", ""),
                    "metric": metric_row.get("metric", ""),
                    "formula_id": formula_id,
                    "official_weight": official_weight,
                    "formula_value_status": metric_accuracy.get("value_status", "not_run" if raw_status != "ok" else ""),
                    "metric_blocking": metric_blocking,
                    "field_name": field_name,
                    "field_category": metadata.get("field_category", ""),
                    "coverage_status": coverage_status,
                    "primary_value": raw_value,
                    "source_sheet": raw_row.get("source_sheet", ""),
                    "source_cell": raw_row.get("source_cell", ""),
                    "source_label": raw_row.get("source_label", ""),
                    "source_type": source_type,
                    "no_cheat_allowed": no_cheat_allowed,
                    "preferred_source": metadata.get("preferred_source", ""),
                    "fallback_source": metadata.get("fallback_source", ""),
                    "likely_sources": metadata.get("likely_sources", ""),
                    "other_page_candidate_statuses": other_evidence["other_page_candidate_statuses"],
                    "other_page_evidence_sheets": other_evidence["other_page_evidence_sheets"],
                    "other_page_official_match": other_evidence["other_page_official_match"],
                    "field_blocking": field_blocking,
                    "suspected_cause": metric_accuracy.get("suspected_cause", raw_status),
                }
            )
    return pd.DataFrame(rows)


def _append_sp_local_support_tab_candidates(
    rows: list[dict[str, Any]],
    case: ScorecardCase,
    primary_values: dict[str, float | None],
    official_values: dict[str, float | None],
) -> None:
    if case.methodology_id != "sp_local_gov_k12":
        return
    support = _load_sp_local_gov_supplemental_report(case.workbook_path)
    if not isinstance(support, pd.DataFrame) or support.empty:
        return
    for _, row in support.iterrows():
        field = str(row.get("field_name", "") or "").strip()
        candidate = _safe_float(row.get("value"))
        primary = primary_values.get(field)
        official = official_values.get(field)
        if primary is None:
            status = "support_no_primary_value"
            delta = None
        elif candidate is None:
            status = "support_missing_value"
            delta = None
        else:
            delta = candidate - primary
            status = "support_primary_match" if abs(delta) <= max(abs(primary) * 0.0001, 0.01) else "support_primary_mismatch"

        official_delta = None if candidate is None or official is None else candidate - official
        official_match = (
            None
            if official_delta is None
            else abs(official_delta) <= max(abs(official or 0) * 0.0001, 0.0001)
        )
        rows.append(
            {
                "fixture_key": case.fixture_key,
                "methodology_id": case.methodology_id,
                "issuer_name": case.issuer_name,
                "workbook": case.workbook_path.name,
                "primary_raw_sheet": case.primary_raw_sheet or "",
                "sheet_name": row.get("sheet_name", ""),
                "field_name": field,
                "matched_label": row.get("matched_label", ""),
                "candidate_cell": row.get("matched_column", ""),
                "candidate_value": candidate,
                "primary_value": primary,
                "delta_to_primary": delta,
                "official_value": official,
                "delta_to_official": official_delta,
                "official_match": official_match,
                "candidate_status": status,
                "match_method": row.get("match_method", ""),
                "notes": row.get("notes", ""),
            }
        )


def _other_page_candidates(case: ScorecardCase, primary_raw: pd.DataFrame, official: pd.DataFrame) -> pd.DataFrame:
    if not case.workbook_path.exists():
        return pd.DataFrame()
    sheet_names = _load_sheet_names(case.workbook_path)
    primary_values = {}
    if not primary_raw.empty:
        primary_values = {
            str(row["field_name"]): _safe_float(row.get("value"))
            for _, row in primary_raw.iterrows()
        }
    official_values = _official_value_lookup(official)

    rows = []
    for sheet_name in sheet_names:
        if sheet_name == case.primary_raw_sheet:
            continue
        try:
            _, report = map_creditscope_workbook(
                uploaded_file=case.workbook_path,
                sheet_name=sheet_name,
                mapping_path="config/field_mapping.csv",
                row_mapping_path="config/creditscope_row_mapping.csv",
            )
        except Exception as exc:
            rows.append(
                {
                    "fixture_key": case.fixture_key,
                    "workbook": case.workbook_path.name,
                    "sheet_name": sheet_name,
                    "field_name": "",
                    "candidate_status": "sheet_parse_error",
                    "notes": str(exc),
                }
            )
            continue
        if not isinstance(report, pd.DataFrame) or report.empty:
            continue
        ready = report[report.get("status", pd.Series(dtype=str)).astype(str).isin(["ready", "review", "missing_value"])].copy()
        if ready.empty:
            continue
        for _, row in ready.iterrows():
            field = str(row.get("field_name", "") or "").strip()
            candidate = _safe_float(row.get("value"))
            primary = primary_values.get(field)
            if primary is None:
                status = "no_primary_value"
                delta = None
            elif candidate is None:
                status = "candidate_missing_value"
                delta = None
            else:
                delta = candidate - primary
                status = "match" if abs(delta) <= max(abs(primary) * 0.0001, 0.01) else "mismatch"
            rows.append(
                {
                    "fixture_key": case.fixture_key,
                    "methodology_id": case.methodology_id,
                    "issuer_name": case.issuer_name,
                    "workbook": case.workbook_path.name,
                    "primary_raw_sheet": case.primary_raw_sheet or "",
                    "sheet_name": sheet_name,
                    "field_name": field,
                    "matched_label": row.get("matched_label", ""),
                    "candidate_cell": row.get("matched_column", ""),
                    "candidate_value": candidate,
                    "primary_value": primary,
                    "delta_to_primary": delta,
                    "official_value": official_values.get(field),
                    "delta_to_official": None
                    if candidate is None or official_values.get(field) is None
                    else candidate - official_values.get(field),
                    "official_match": None
                    if candidate is None or official_values.get(field) is None
                    else abs(candidate - official_values.get(field)) <= max(abs(official_values.get(field) or 0) * 0.0001, 0.0001),
                    "candidate_status": status,
                    "match_method": row.get("match_method", ""),
                    "notes": row.get("notes", ""),
                }
            )
    _append_sp_local_support_tab_candidates(rows, case, primary_values, official_values)
    return pd.DataFrame(rows)


def build_accuracy_package(output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    summary_rows: list[dict[str, Any]] = []
    accuracy_frames: list[pd.DataFrame] = []
    raw_frames: list[pd.DataFrame] = []
    raw_report_frames: list[pd.DataFrame] = []
    page_frames: list[pd.DataFrame] = []
    coverage_frames: list[pd.DataFrame] = []
    source_quality_frames: list[pd.DataFrame] = []

    for case in CASES:
        official_path = Path("config/validation_fixtures") / f"{case.fixture_key}.csv"
        official = load_official_fixture(official_path) if official_path.exists() else pd.DataFrame()
        primary_raw, raw_report, raw_status = _raw_fixture_from_primary(case)

        if not primary_raw.empty:
            formula_results = calculate_template_formulas_from_raw(primary_raw)
            value_comparison = compare_raw_formula_values_to_official(formula_results, official, primary_raw)
            rating_output, _ = _run_rating(case, formula_results)
            score_comparison = compare_auto_scores_to_official(rating_output, official)
            source_quality = raw_source_quality_summary(primary_raw)
        else:
            formula_results = pd.DataFrame()
            value_comparison = pd.DataFrame()
            rating_output = {}
            score_comparison = pd.DataFrame()
            source_quality = pd.DataFrame()

        accuracy = _accuracy_rows(case, official, value_comparison, score_comparison, raw_status)
        if not accuracy.empty:
            accuracy_frames.append(accuracy)
        pages = _other_page_candidates(case, primary_raw, official)
        if not pages.empty:
            page_frames.append(pages)
        coverage = _field_coverage_rows(case, official, primary_raw, raw_status, accuracy, pages)
        if not coverage.empty:
            coverage_frames.append(coverage)
        if not primary_raw.empty:
            raw_frames.append(primary_raw)
        if isinstance(raw_report, pd.DataFrame) and not raw_report.empty:
            report = raw_report.copy()
            report.insert(0, "fixture_key", case.fixture_key)
            raw_report_frames.append(report)
        if not source_quality.empty:
            source_quality = source_quality.copy()
            source_quality.insert(0, "fixture_key", case.fixture_key)
            source_quality_frames.append(source_quality)

        value_counts = _source_summary_for_compare(value_comparison)
        score_counts = _status_counts(score_comparison, "score_match")
        rating = _rating_result(rating_output)
        official_first = official.iloc[0] if not official.empty else {}
        summary_rows.append(
            {
                "fixture_key": case.fixture_key,
                "methodology_id": case.methodology_id,
                "issuer_name": case.issuer_name,
                "workbook": case.workbook_path.name,
                "primary_raw_sheet": case.primary_raw_sheet or "",
                "raw_status": raw_status,
                "primary_input_fields": len(primary_raw),
                "formula_rows": len(formula_results),
                "official_metric_rows": len(official),
                "value_matches": value_counts.get("match", 0),
                "value_mismatches": value_counts.get("mismatch", 0),
                "model_missing": value_counts.get("model_missing", 0),
                "manual_skips": value_counts.get("manual_skip", 0),
                "score_true": score_counts.get("True", 0),
                "score_false": score_counts.get("False", 0),
                "model_rating": rating.get("indicative_rating", ""),
                "official_rating": official_first.get("official_rating", "") if hasattr(official_first, "get") else "",
                "model_weighted_score": rating.get("overall_score", ""),
                "official_weighted_score": official_first.get("official_weighted_score", "") if hasattr(official_first, "get") else "",
                "coverage_status": rating.get("coverage_status", ""),
                "notes": case.notes,
            }
        )

    outputs = {
        "summary": output_dir / "summary.csv",
        "accuracy_matrix": output_dir / "accuracy_matrix.csv",
        "primary_raw_inputs": output_dir / "primary_raw_inputs.csv",
        "raw_mapping_report": output_dir / "raw_mapping_report.csv",
        "page_consistency": output_dir / "page_consistency.csv",
        "field_coverage": output_dir / "field_coverage.csv",
        "source_quality": output_dir / "source_quality.csv",
        "tables_json": output_dir / "tables.json",
        "manifest": output_dir / "manifest.json",
    }
    tables = {
        "summary": pd.DataFrame(summary_rows),
        "accuracy_matrix": pd.concat(accuracy_frames, ignore_index=True) if accuracy_frames else pd.DataFrame(),
        "primary_raw_inputs": pd.concat(raw_frames, ignore_index=True) if raw_frames else pd.DataFrame(),
        "raw_mapping_report": pd.concat(raw_report_frames, ignore_index=True) if raw_report_frames else pd.DataFrame(),
        "page_consistency": pd.concat(page_frames, ignore_index=True) if page_frames else pd.DataFrame(),
        "field_coverage": pd.concat(coverage_frames, ignore_index=True) if coverage_frames else pd.DataFrame(),
        "source_quality": pd.concat(source_quality_frames, ignore_index=True) if source_quality_frames else pd.DataFrame(),
    }
    for name, frame in tables.items():
        frame.to_csv(outputs[name], index=False)
    outputs["tables_json"].write_text(
        json.dumps(
            {
                name: _records_for_json(frame)
                for name, frame in tables.items()
            },
            indent=2,
            allow_nan=False,
            default=_json_default,
        ),
        encoding="utf-8",
    )
    manifest = {
        "created_by": "scripts/methodology_accuracy_matrix.py",
        "rule": "Primary model inputs are sourced only from configured CreditScope/raw tabs; official scorecard tabs are benchmark/evidence only.",
        "cases": [case.__dict__ | {"workbook_path": str(case.workbook_path)} for case in CASES],
        "outputs": {name: str(path) for name, path in outputs.items() if name != "manifest"},
        "row_counts": {name: int(len(frame)) for name, frame in tables.items()},
    }
    outputs["manifest"].write_text(json.dumps(manifest, indent=2, default=_json_default), encoding="utf-8")
    return outputs


def main() -> None:
    parser = argparse.ArgumentParser(description="Build no-cheat methodology accuracy matrix source tables.")
    parser.add_argument(
        "--output-dir",
        default="work/methodology_accuracy_matrix",
        help="Directory for generated CSV/JSON tables.",
    )
    args = parser.parse_args()
    outputs = build_accuracy_package(Path(args.output_dir))
    print(json.dumps({name: str(path) for name, path in outputs.items()}, indent=2))


if __name__ == "__main__":
    main()
