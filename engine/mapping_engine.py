"""
Mapping Engine for CreditScope / IPEDS / OS / ACFR uploads.

Purpose
-------
Convert uploaded source files with human-facing column names into canonical
issuer_data fields used by the formula engine.

Example
-------
CreditScope CSV columns:
    Population, Full Value, Net Direct Debt

field_mapping.csv:
    field_name,source_name,possible_column_names,match_type,notes
    population,CreditScope,"Population|Resident Population",fuzzy,...

Output:
    issuer_data = {
        "population": 530000,
        "full_value": 40000000000,
        "net_direct_debt": 700000000,
    }
"""

from __future__ import annotations

from dataclasses import dataclass, asdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_MAPPING_PATH = Path("config/field_mapping.csv")
DEFAULT_CREDITSCOPE_ROW_MAPPING_PATH = Path("config/creditscope_row_mapping.csv")

CREDITSCOPE_THOUSAND_DOLLAR_FIELDS = {
    "full_value",
    "assessed_value",
    "assessed_value_prior",
    "operating_revenue",
    "operating_expense",
    "governmental_revenue",
    "governmental_expense",
    "transfers",
    "committed_fund_balance",
    "assigned_fund_balance",
    "unassigned_fund_balance",
    "fund_balance",
    "cash",
    "cash_and_investments",
    "net_assets",
    "revenue",
    "debt",
    "long_term_debt",
    "net_direct_debt",
    "debt_service",
    "pension_cost",
    "opeb_cost",
    "net_pension_liability",
    "mads",
    "adjusted_npl",
    "adjusted_opeb",
}

CREDITSCOPE_SERIES_FIELDS = {
    "governmental_revenue",
    "governmental_expense",
    "operating_revenue",
    "operating_expense",
    "operating_transfers",
    "committed_fund_balance",
    "assigned_fund_balance",
    "unassigned_fund_balance",
    "reserve_revenue",
}

CREDITSCOPE_SERIES_PERIODS = 3

DERIVED_CREDITSCOPE_LABEL_PATTERNS = [
    r"\bper capita\b",
    r"\bratio\b",
    r"\bmargin\b",
    r"\bburden\b",
    r"\bcoverage\b",
    r"\bdays\b",
    r"\bscore\b",
    r"\brating\b",
    r"\baverage\b",
    r"\bavg\b",
    r"\btrend\b",
    r"\bchange\b",
    r"\bas %\b",
    r"\bpercent\b",
    r"\bpct\b",
    r"\bto full value\b",
    r"\bto revenue\b",
    r"\bto debt\b",
    r"\bx$",
]

STRICT_EXPLICIT_FIELDS = {
    "service_area_population",
    "cash_and_investments",
    "long_term_debt",
    "net_direct_debt",
    "debt_service",
    "pension_cost",
    "opeb_cost",
    "net_pension_liability",
}


@dataclass
class FieldMatch:
    """One canonical field mapped from one uploaded column."""

    field_name: str
    source_name: str
    matched_column: Optional[str]
    matched_label: Optional[str]
    match_method: str  # exact_alias, normalized_alias, fuzzy_alias, not_found
    confidence: float
    value: Any = None
    notes: str = ""

    @property
    def status(self) -> str:
        if self.matched_column is None:
            return "missing"
        if self.value is None or (isinstance(self.value, float) and pd.isna(self.value)):
            return "missing_value"
        if self.confidence >= 0.97:
            return "ready"
        if self.confidence >= 0.78:
            return "review"
        return "low_confidence"


def normalize_label(value: Any) -> str:
    """Normalize labels for resilient matching."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def is_likely_derived_creditscope_label(label: Any) -> bool:
    """Return True for labels that look like CreditScope-calculated metrics."""
    norm = normalize_label(label)
    if not norm:
        return False
    return any(re.search(pattern, norm) for pattern in DERIVED_CREDITSCOPE_LABEL_PATTERNS)


def split_aliases(value: Any) -> List[str]:
    """Split pipe-delimited aliases from field_mapping.csv."""
    if pd.isna(value):
        return []
    return [x.strip() for x in str(value).split("|") if x.strip()]


def load_mapping_table(mapping_path: Union[str, Path] = DEFAULT_MAPPING_PATH) -> pd.DataFrame:
    """Load and validate config/field_mapping.csv."""
    path = Path(mapping_path)
    if not path.exists():
        raise FileNotFoundError(f"Mapping file not found: {path}")

    mapping = pd.read_csv(path)
    required = {"field_name", "source_name", "possible_column_names"}
    missing = required - set(mapping.columns)
    if missing:
        raise ValueError(f"field_mapping.csv missing required columns: {sorted(missing)}")

    mapping = mapping.copy()
    mapping["field_name"] = mapping["field_name"].astype(str).str.strip()
    mapping["source_name"] = mapping["source_name"].astype(str).str.strip()
    mapping["possible_column_names"] = mapping["possible_column_names"].fillna("")
    if "match_type" not in mapping.columns:
        mapping["match_type"] = "fuzzy"
    if "notes" not in mapping.columns:
        mapping["notes"] = ""
    return mapping


def read_uploaded_file(uploaded_file: Any) -> pd.DataFrame:
    """
    Read csv/xlsx from either a Streamlit UploadedFile or a local path.
    """
    if isinstance(uploaded_file, (str, Path)):
        filename = str(uploaded_file)
        suffix = Path(filename).suffix.lower()
    else:
        filename = getattr(uploaded_file, "name", "")
        suffix = Path(filename).suffix.lower()

    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(uploaded_file)
    if suffix == ".csv":
        return pd.read_csv(uploaded_file)
    raise ValueError(f"Unsupported file type: {suffix or filename}. Use CSV or Excel.")


def _open_workbook(uploaded_file: Any, data_only: bool = True):
    """Open a local path or Streamlit UploadedFile with openpyxl."""
    if not isinstance(uploaded_file, (str, Path)) and hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    return load_workbook(uploaded_file, read_only=True, data_only=data_only)


def _sheet_raw_label_score(ws: Any, max_rows: int = 1200) -> int:
    raw_terms = {
        "full value",
        "total assessed value",
        "population",
        "operating revenue",
        "operating expenses",
        "revenue",
        "total debt",
        "cash and short term investments",
        "net assets",
        "net position",
        "fund balance",
    }
    score = 0
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        label = normalize_label(ws.cell(row_idx, 1).value)
        if not label or is_likely_derived_creditscope_label(label):
            continue
        if label in raw_terms:
            score += 3
        elif any(term in label for term in raw_terms):
            score += 1
    return score


def _pick_creditscope_sheet(workbook: Any, sheet_name: Optional[str] = None):
    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    scored = []
    for name in workbook.sheetnames:
        norm_name = normalize_label(name)
        ws = workbook[name]
        score = _sheet_raw_label_score(ws)
        if "credit scope" in norm_name or "creditscope" in norm_name:
            score += 5
        if "fin" in norm_name or "financial" in norm_name:
            score += 3
        if norm_name in {"public", "scorecard"}:
            score -= 8
        scored.append((score, name))
    scored.sort(reverse=True)
    if scored and scored[0][0] > 0:
        return workbook[scored[0][1]]

    preferred_terms = ["credit scope", "creditscope", "fin", "financial", "contra costa"]
    for term in preferred_terms:
        for name in workbook.sheetnames:
            if term in normalize_label(name):
                return workbook[name]
    return workbook[workbook.sheetnames[0]]


def _sheet_label_value_frame(ws: Any, value_col: int = 2, max_rows: int = 1200) -> pd.DataFrame:
    rows = []
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        label = ws.cell(row_idx, 1).value
        value = ws.cell(row_idx, value_col).value
        if label is None:
            continue
        label_text = str(label).strip()
        if not label_text:
            continue
        rows.append(
            {
                "row": row_idx,
                "label": label_text,
                "normalized_label": normalize_label(label_text),
                "value": clean_numeric_value(value),
            }
        )
    return pd.DataFrame(rows)


def _scale_creditscope_value(field_name: str, value: Any) -> Any:
    """CreditScope workbook financial statement values are usually shown in $000s."""
    if isinstance(value, (list, tuple)):
        return [_scale_creditscope_value(field_name, item) for item in value]
    if field_name not in CREDITSCOPE_THOUSAND_DOLLAR_FIELDS:
        return value
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return value
    if isinstance(value, (int, float)):
        return float(value) * 1000.0
    return value


def _scale_raw_value(value: Any, scale: float) -> Any:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return value
    if isinstance(value, (int, float)):
        return float(value) * float(scale)
    return value


def _read_creditscope_series(
    ws: Any,
    *,
    row_number: int,
    start_col: int,
    scale: float,
    periods: int = CREDITSCOPE_SERIES_PERIODS,
) -> tuple[Optional[list[Any]], Optional[str]]:
    values: list[Any] = []
    first_col: Optional[int] = None
    last_col: Optional[int] = None
    for col_idx in range(start_col, start_col + periods):
        value = clean_numeric_value(ws.cell(row_number, col_idx).value)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            continue
        value = _scale_raw_value(value, scale)
        values.append(value)
        first_col = col_idx if first_col is None else first_col
        last_col = col_idx

    if len(values) < 2 or first_col is None or last_col is None:
        return None, None
    cell_range = f"{ws.title.strip()}!{get_column_letter(first_col)}{row_number}:{get_column_letter(last_col)}{row_number}"
    return values, cell_range


def _load_creditscope_row_mapping(path: Union[str, Path] = DEFAULT_CREDITSCOPE_ROW_MAPPING_PATH) -> pd.DataFrame:
    path = Path(path)
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_csv(path)
    required = {"field_name", "exact_label", "value_col_offset", "scale"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{path} missing required columns: {sorted(missing)}")
    df = df.copy()
    df["field_name"] = df["field_name"].astype(str).str.strip()
    df["exact_label"] = df["exact_label"].astype(str).str.strip()
    df["value_col_offset"] = pd.to_numeric(df["value_col_offset"], errors="coerce").fillna(0).astype(int)
    df["scale"] = pd.to_numeric(df["scale"], errors="coerce").fillna(1.0)
    return df


def _best_label_match(
    field_name: str,
    aliases: Iterable[str],
    labels: Iterable[str],
    fuzzy_threshold: float = 0.92,
) -> Tuple[Optional[str], Optional[str], str, float]:
    labels = [label for label in labels if not is_likely_derived_creditscope_label(label)]
    aliases = list(aliases)
    if field_name in STRICT_EXPLICIT_FIELDS:
        aliases = [alias for alias in aliases if normalize_label(alias) != normalize_label(field_name.replace("_", " "))]

    normalized_labels = {normalize_label(label): label for label in labels}
    for alias in aliases:
        norm_alias = normalize_label(alias)
        if norm_alias in normalized_labels:
            return normalized_labels[norm_alias], alias, "normalized_label", 0.97

    best: Tuple[Optional[str], Optional[str], float] = (None, None, 0.0)
    for alias in aliases:
        norm_alias = normalize_label(alias)
        if not norm_alias:
            continue
        for label in labels:
            norm_label = normalize_label(label)
            score = SequenceMatcher(None, norm_alias, norm_label).ratio()
            if score > best[2]:
                best = (label, alias, score)

    if best[0] is not None and best[2] >= fuzzy_threshold:
        return best[0], best[1], "fuzzy_label", round(best[2], 3)
    return None, None, "not_found", 0.0


def map_creditscope_workbook(
    uploaded_file: Any,
    mapping_path: Union[str, Path] = DEFAULT_MAPPING_PATH,
    row_mapping_path: Union[str, Path] = DEFAULT_CREDITSCOPE_ROW_MAPPING_PATH,
    sheet_name: Optional[str] = None,
    value_col: int = 2,
    fuzzy_threshold: float = 0.92,
) -> Tuple[Dict[str, Any], pd.DataFrame]:
    """
    Map CreditScope-style workbooks where column A contains metric labels and
    each following column is a reporting period.

    The Ramirez reference files use this shape: row labels in column A and
    current-year values in column B. This loader is intentionally separate from
    the simple tabular column matcher so both layouts can coexist.
    """
    workbook = _open_workbook(uploaded_file, data_only=True)
    ws = _pick_creditscope_sheet(workbook, sheet_name=sheet_name)
    label_values = _sheet_label_value_frame(ws, value_col=value_col)
    mapping = load_mapping_table(mapping_path)
    source_rows = mapping[mapping["source_name"].map(normalize_label) == normalize_label("CreditScope")]
    exact_mapping = _load_creditscope_row_mapping(row_mapping_path)

    issuer_data: Dict[str, Any] = {}
    matches: List[FieldMatch] = []
    labels = label_values["label"].tolist() if not label_values.empty else []
    sheet_title = ws.title.strip()
    if not label_values.empty:
        label_lookup_df = label_values.drop_duplicates("normalized_label", keep="first")
        by_norm_label = label_lookup_df.set_index("normalized_label").to_dict(orient="index")
    else:
        by_norm_label = {}

    for _, row in source_rows.iterrows():
        field_name = str(row["field_name"]).strip()
        aliases = split_aliases(row.get("possible_column_names", ""))
        notes = "" if pd.isna(row.get("notes", "")) else str(row.get("notes", ""))

        matched_label = None
        matched_alias = None
        matched_cell = None
        method = "not_found"
        confidence = 0.0
        value = None

        exact_rows = exact_mapping[exact_mapping["field_name"] == field_name] if not exact_mapping.empty else pd.DataFrame()
        for _, exact in exact_rows.iterrows():
            norm_exact = normalize_label(exact["exact_label"])
            if norm_exact not in by_norm_label:
                continue
            row_number = int(by_norm_label[norm_exact]["row"])
            target_col = int(value_col) + int(exact["value_col_offset"])
            scale = float(exact["scale"])
            candidate_cell = f"{sheet_title}!{get_column_letter(target_col)}{row_number}"
            candidate_method = "exact_row_mapping"

            if field_name in CREDITSCOPE_SERIES_FIELDS:
                series_value, series_cell = _read_creditscope_series(
                    ws,
                    row_number=row_number,
                    start_col=target_col,
                    scale=scale,
                )
                if series_value is not None:
                    value = series_value
                    candidate_cell = series_cell
                    candidate_method = "exact_row_mapping_series"
                else:
                    value = clean_numeric_value(ws.cell(row_number, target_col).value)
                    value = _scale_raw_value(value, scale)
            else:
                value = clean_numeric_value(ws.cell(row_number, target_col).value)
                value = _scale_raw_value(value, scale)

            if value is None or (isinstance(value, float) and pd.isna(value)):
                continue
            matched_label = str(by_norm_label[norm_exact]["label"])
            matched_cell = candidate_cell
            method = candidate_method
            confidence = 1.0
            notes = str(exact.get("notes", notes) or notes)
            break

        if matched_label is None:
            matched_label, matched_alias, method, confidence = _best_label_match(
                field_name=field_name,
                aliases=aliases,
                labels=labels,
                fuzzy_threshold=fuzzy_threshold,
            )
            matched_row = by_norm_label.get(normalize_label(matched_label)) if matched_label is not None else None
            value = matched_row.get("value") if matched_row else None
            if matched_row:
                row_number = int(matched_row["row"])
                matched_cell = f"{sheet_title}!{get_column_letter(value_col)}{row_number}"
                if field_name in CREDITSCOPE_SERIES_FIELDS:
                    scale = 1000.0 if field_name in CREDITSCOPE_THOUSAND_DOLLAR_FIELDS else 1.0
                    series_value, series_cell = _read_creditscope_series(
                        ws,
                        row_number=row_number,
                        start_col=int(value_col),
                        scale=scale,
                    )
                    if series_value is not None:
                        value = series_value
                        matched_cell = series_cell
                        method = f"{method}_series"
                    else:
                        value = _scale_creditscope_value(field_name, value)
                else:
                    value = _scale_creditscope_value(field_name, value)

        if matched_label is not None and value is not None and not (isinstance(value, float) and pd.isna(value)):
            issuer_data[field_name] = value
        matches.append(
            FieldMatch(
                field_name=field_name,
                source_name="CreditScope",
                matched_column=matched_cell if matched_label is not None else None,
                matched_label=matched_label or matched_alias,
                match_method=method,
                confidence=confidence,
                value=value,
                notes=notes,
            )
        )

    match_report = pd.DataFrame([asdict(m) | {"status": m.status} for m in matches])
    if not match_report.empty:
        match_report.insert(0, "sheet_name", ws.title)
    return issuer_data, match_report


def _best_column_match(
    aliases: Iterable[str],
    columns: Iterable[str],
    fuzzy_threshold: float = 0.82,
) -> Tuple[Optional[str], Optional[str], str, float]:
    """
    Find best uploaded column for a list of aliases.

    Returns:
        matched_column, matched_alias, match_method, confidence
    """
    columns = list(columns)
    aliases = list(aliases)

    # 1. Exact raw string match
    for alias in aliases:
        for col in columns:
            if str(col).strip() == alias.strip():
                return col, alias, "exact_alias", 1.0

    # 2. Normalized exact match
    normalized_cols = {normalize_label(col): col for col in columns}
    for alias in aliases:
        norm_alias = normalize_label(alias)
        if norm_alias in normalized_cols:
            return normalized_cols[norm_alias], alias, "normalized_alias", 0.97

    # 3. Fuzzy match against normalized strings
    best: Tuple[Optional[str], Optional[str], float] = (None, None, 0.0)
    for alias in aliases:
        norm_alias = normalize_label(alias)
        if not norm_alias:
            continue
        for col in columns:
            score = SequenceMatcher(None, norm_alias, normalize_label(col)).ratio()
            if score > best[2]:
                best = (col, alias, score)

    if best[0] is not None and best[2] >= fuzzy_threshold:
        return best[0], best[1], "fuzzy_alias", round(best[2], 3)

    return None, None, "not_found", 0.0


def clean_numeric_value(value: Any) -> Any:
    """
    Convert common financial strings to numbers when safe.

    Handles:
        "$1,234" -> 1234
        "12.5%" -> 0.125
        "(1,234)" -> -1234
        "N/A" -> None
    Leaves non-numeric strings unchanged.
    """
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return value

    raw = str(value).strip()
    if raw == "" or raw.lower() in {"na", "n/a", "none", "null", "--", "-"}:
        return None

    is_percent = raw.endswith("%")
    negative = raw.startswith("(") and raw.endswith(")")

    text = raw.replace("$", "").replace(",", "").replace("%", "")
    text = text.replace("(", "").replace(")", "").strip()

    try:
        number = float(text)
        if negative:
            number = -number
        if is_percent:
            number = number / 100.0
        return number
    except ValueError:
        return raw


def extract_value_from_column(df: pd.DataFrame, column: str, row_index: int = 0) -> Any:
    """
    Extract one value from a matched column.

    MVP assumption: one issuer / one period per uploaded file, so the first
    non-empty value is usually the target value. If that fails, fall back to
    the requested row_index.
    """
    series = df[column]
    non_empty = series.dropna()
    if len(non_empty) > 0:
        return clean_numeric_value(non_empty.iloc[row_index if row_index < len(non_empty) else 0])
    return None


def map_uploaded_dataframe(
    df: pd.DataFrame,
    source_name: str,
    mapping: Union[pd.DataFrame, str, Path] = DEFAULT_MAPPING_PATH,
    fuzzy_threshold: float = 0.82,
    row_index: int = 0,
) -> Tuple[Dict[str, Any], pd.DataFrame]:
    """
    Map uploaded dataframe columns to canonical issuer_data.

    Args:
        df: uploaded CreditScope / IPEDS / ACFR / OS dataframe.
        source_name: source label, e.g. "CreditScope", "IPEDS", "ACFR".
        mapping: mapping dataframe or path to field_mapping.csv.
        fuzzy_threshold: minimum fuzzy score accepted.
        row_index: row index used when multiple non-empty rows exist.

    Returns:
        issuer_data: dict keyed by canonical field_name.
        match_report: dataframe with status, matched column, confidence, value.
    """
    if not isinstance(mapping, pd.DataFrame):
        mapping_df = load_mapping_table(mapping)
    else:
        mapping_df = mapping.copy()

    source_norm = normalize_label(source_name)
    source_rows = mapping_df[mapping_df["source_name"].map(normalize_label) == source_norm]

    # Allow generic/common mapping rows if you later add source_name = Any/Common.
    generic_rows = mapping_df[mapping_df["source_name"].map(normalize_label).isin({"any", "common", "all"})]
    source_rows = pd.concat([source_rows, generic_rows], ignore_index=True).drop_duplicates(
        subset=["field_name", "possible_column_names"], keep="first"
    )

    issuer_data: Dict[str, Any] = {}
    matches: List[FieldMatch] = []

    for _, row in source_rows.iterrows():
        field_name = str(row["field_name"]).strip()
        aliases = split_aliases(row.get("possible_column_names", ""))
        notes = "" if pd.isna(row.get("notes", "")) else str(row.get("notes", ""))

        matched_col, matched_alias, method, confidence = _best_column_match(
            aliases=aliases,
            columns=df.columns,
            fuzzy_threshold=fuzzy_threshold,
        )

        value = None
        if matched_col is not None:
            value = extract_value_from_column(df, matched_col, row_index=row_index)
            if value is not None and not (isinstance(value, float) and pd.isna(value)):
                issuer_data[field_name] = value

        matches.append(
            FieldMatch(
                field_name=field_name,
                source_name=source_name,
                matched_column=matched_col,
                matched_label=matched_alias,
                match_method=method,
                confidence=confidence,
                value=value,
                notes=notes,
            )
        )

    match_report = pd.DataFrame([asdict(m) | {"status": m.status} for m in matches])
    if not match_report.empty:
        match_report = match_report[
            [
                "field_name",
                "source_name",
                "status",
                "matched_column",
                "matched_label",
                "match_method",
                "confidence",
                "value",
                "notes",
            ]
        ]
    return issuer_data, match_report


def map_uploaded_file(
    uploaded_file: Any,
    source_name: str,
    mapping_path: Union[str, Path] = DEFAULT_MAPPING_PATH,
    fuzzy_threshold: float = 0.82,
    row_index: int = 0,
) -> Tuple[Dict[str, Any], pd.DataFrame]:
    """
    Convenience wrapper for Streamlit uploaders.
    """
    df = read_uploaded_file(uploaded_file)
    return map_uploaded_dataframe(
        df=df,
        source_name=source_name,
        mapping=mapping_path,
        fuzzy_threshold=fuzzy_threshold,
        row_index=row_index,
    )


def merge_issuer_data(
    *issuer_data_dicts: Dict[str, Any],
    overwrite: bool = False,
) -> Dict[str, Any]:
    """
    Merge mapped outputs from multiple sources.

    By default, earlier sources win. This is useful when you want source priority:
        CreditScope first, then IPEDS, then ACFR, then manual.
    Set overwrite=True if later sources should replace earlier values.
    """
    merged: Dict[str, Any] = {}
    for data in issuer_data_dicts:
        for key, value in data.items():
            if overwrite or key not in merged or merged[key] is None:
                merged[key] = value
    return merged


if __name__ == "__main__":
    # Tiny smoke test. Run from project root:
    # python engine/mapping_engine.py
    sample = pd.DataFrame(
        {
            "Population": [530000],
            "Full Value": ["$40,000,000,000"],
            "Net Direct Debt": ["700,000,000"],
        }
    )
    data, report = map_uploaded_dataframe(sample, source_name="CreditScope")
    print(data)
    print(report.head(10).to_string(index=False))
